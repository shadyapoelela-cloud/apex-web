# APEX Analyzer v2.1 - Updated 2026-03-27
"""
APEX Financial Analyzer — Core Engine
أبيكس للاستشارات المالية والاستثمارية

يحسب 16+ نسبة مالية تلقائياً من ملف Excel أو PDF
ويولّد درجة الجاهزية الاستثمارية من 100

التثبيت:
    pip install pandas openpyxl pdfplumber openai fastapi uvicorn python-multipart

الاستخدام:
    python apex_analyzer.py --file financial_statements.xlsx
    python apex_analyzer.py --file financial_statements.xlsx --openai-key sk-...
"""

import json
import argparse
from dataclasses import dataclass, field, asdict
from typing import Optional

# ─── DATA MODELS ──────────────────────────────────────────────────────────────

@dataclass
class FinancialData:
    """البيانات المالية المستخرجة من القوائم المالية"""
    
    # قائمة الدخل
    revenue: float = 0                    # الإيرادات
    cost_of_goods_sold: float = 0         # تكلفة المبيعات
    gross_profit: float = 0              # الربح الإجمالي
    operating_expenses: float = 0        # المصاريف التشغيلية
    ebitda: float = 0                    # الأرباح قبل الفوائد والضرائب والاستهلاك
    ebit: float = 0                      # الأرباح قبل الفوائد والضرائب
    interest_expense: float = 0          # مصاريف الفوائد
    net_profit: float = 0               # صافي الربح
    
    # الميزانية العمومية
    current_assets: float = 0           # الأصول المتداولة
    cash: float = 0                     # النقدية
    accounts_receivable: float = 0      # حسابات القبض
    inventory: float = 0               # المخزون
    total_assets: float = 0            # إجمالي الأصول
    current_liabilities: float = 0     # الالتزامات المتداولة
    total_liabilities: float = 0       # إجمالي الالتزامات
    equity: float = 0                  # حقوق الملكية
    long_term_debt: float = 0         # الديون طويلة الأجل
    
    # بيانات إضافية
    days_in_period: int = 365          # عدد أيام الفترة
    previous_revenue: float = 0        # إيرادات الفترة السابقة
    industry: str = "general"          # القطاع


@dataclass
class FinancialRatio:
    """نسبة مالية واحدة مع تفسيرها"""
    name_ar: str
    name_en: str
    value: float
    unit: str                           # %, x, days
    benchmark: float                    # المعيار القطاعي
    score: float                        # الدرجة من 100
    interpretation: str                 # التفسير بالعربية
    status: str                         # good / warning / danger


@dataclass
class AnalysisResult:
    """نتيجة التحليل الكاملة"""
    company_name: str = "شركتك"
    period: str = "2024"
    ratios: list = field(default_factory=list)
    readiness_score: float = 0
    readiness_label: str = ""
    ai_insights: list = field(default_factory=list)
    recommendations: list = field(default_factory=list)
    forecast: dict = field(default_factory=dict)


# ─── INDUSTRY BENCHMARKS (السعودية) ──────────────────────────────────────────

BENCHMARKS = {
    "general": {
        "gross_margin": 35.0,
        "net_margin": 8.0,
        "ebitda_margin": 15.0,
        "roe": 12.0,
        "roa": 6.0,
        "current_ratio": 1.5,
        "quick_ratio": 1.0,
        "debt_to_equity": 1.0,
        "interest_coverage": 3.0,
        "asset_turnover": 1.0,
        "inventory_days": 60,
        "receivable_days": 45,
    },
    "retail": {
        "gross_margin": 28.0,
        "net_margin": 4.0,
        "ebitda_margin": 10.0,
        "roe": 15.0,
        "roa": 8.0,
        "current_ratio": 1.2,
        "quick_ratio": 0.7,
        "debt_to_equity": 1.5,
        "interest_coverage": 2.5,
        "asset_turnover": 1.8,
        "inventory_days": 45,
        "receivable_days": 20,
    },
    "manufacturing": {
        "gross_margin": 30.0,
        "net_margin": 7.0,
        "ebitda_margin": 18.0,
        "roe": 13.0,
        "roa": 5.0,
        "current_ratio": 1.8,
        "quick_ratio": 1.0,
        "debt_to_equity": 1.2,
        "interest_coverage": 4.0,
        "asset_turnover": 0.8,
        "inventory_days": 90,
        "receivable_days": 60,
    },
    "services": {
        "gross_margin": 55.0,
        "net_margin": 12.0,
        "ebitda_margin": 22.0,
        "roe": 20.0,
        "roa": 10.0,
        "current_ratio": 1.3,
        "quick_ratio": 1.2,
        "debt_to_equity": 0.5,
        "interest_coverage": 5.0,
        "asset_turnover": 1.5,
        "inventory_days": 10,
        "receivable_days": 35,
    },
}


# ─── CORE CALCULATOR ──────────────────────────────────────────────────────────

class ApexAnalyzer:
    """محرك التحليل المالي الرئيسي"""

    def __init__(self, data: FinancialData, openai_key: Optional[str] = None):
        self.data = data
        self.openai_key = openai_key
        self.bench = BENCHMARKS.get(data.industry, BENCHMARKS["general"])
        self.d = data.days_in_period

    def _score(self, value: float, benchmark: float, higher_is_better: bool = True, max_score: float = 100) -> float:
        """يحسب الدرجة من 100 بناءً على المعيار القطاعي"""
        if benchmark == 0:
            return 50.0
        ratio = value / benchmark
        if higher_is_better:
            score = min(ratio * 70, max_score)
        else:
            score = min((benchmark / max(value, 0.001)) * 70, max_score)
        return round(max(0, score), 1)

    def _status(self, score: float) -> str:
        if score >= 65:
            return "good"
        elif score >= 40:
            return "warning"
        return "danger"

    def calculate_all(self) -> list:
        """يحسب جميع النسب المالية الـ 16+"""
        d = self.data
        b = self.bench
        ratios = []

        # ─── نسب الربحية ──────────────────────────────────────────────────────

        # 1. هامش الربح الإجمالي
        gross_margin = (d.gross_profit / d.revenue * 100) if d.revenue else 0
        s = self._score(gross_margin, b["gross_margin"])
        ratios.append(FinancialRatio(
            name_ar="هامش الربح الإجمالي",
            name_en="Gross Profit Margin",
            value=round(gross_margin, 2),
            unit="%",
            benchmark=b["gross_margin"],
            score=s,
            interpretation=f"مقابل {b['gross_margin']}% معيار القطاع — {'ممتاز' if s >= 65 else 'يحتاج تحسين'}",
            status=self._status(s)
        ))

        # 2. هامش صافي الربح
        net_margin = (d.net_profit / d.revenue * 100) if d.revenue else 0
        s = self._score(net_margin, b["net_margin"])
        ratios.append(FinancialRatio(
            name_ar="هامش صافي الربح",
            name_en="Net Profit Margin",
            value=round(net_margin, 2),
            unit="%",
            benchmark=b["net_margin"],
            score=s,
            interpretation=f"مقابل {b['net_margin']}% معيار القطاع — {'جيد' if s >= 65 else 'منخفض نسبياً'}",
            status=self._status(s)
        ))

        # 3. هامش EBITDA
        ebitda_margin = (d.ebitda / d.revenue * 100) if d.revenue else 0
        s = self._score(ebitda_margin, b["ebitda_margin"])
        ratios.append(FinancialRatio(
            name_ar="هامش EBITDA",
            name_en="EBITDA Margin",
            value=round(ebitda_margin, 2),
            unit="%",
            benchmark=b["ebitda_margin"],
            score=s,
            interpretation=f"مقياس التدفق النقدي التشغيلي — {'قوي' if s >= 65 else 'متوسط'}",
            status=self._status(s)
        ))

        # 4. العائد على حقوق الملكية (ROE)
        roe = (d.net_profit / d.equity * 100) if d.equity else 0
        s = self._score(roe, b["roe"])
        ratios.append(FinancialRatio(
            name_ar="العائد على حقوق الملكية (ROE)",
            name_en="Return on Equity",
            value=round(roe, 2),
            unit="%",
            benchmark=b["roe"],
            score=s,
            interpretation=f"كل 100 ريال استثمار يولّد {roe:.1f} ريال ربح",
            status=self._status(s)
        ))

        # 5. العائد على الأصول (ROA)
        roa = (d.net_profit / d.total_assets * 100) if d.total_assets else 0
        s = self._score(roa, b["roa"])
        ratios.append(FinancialRatio(
            name_ar="العائد على الأصول (ROA)",
            name_en="Return on Assets",
            value=round(roa, 2),
            unit="%",
            benchmark=b["roa"],
            score=s,
            interpretation=f"كفاءة توظيف الأصول في توليد الأرباح",
            status=self._status(s)
        ))

        # ─── نسب السيولة ──────────────────────────────────────────────────────

        # 6. نسبة السيولة الجارية
        current_ratio = (d.current_assets / d.current_liabilities) if d.current_liabilities else 0
        s = self._score(current_ratio, b["current_ratio"])
        ratios.append(FinancialRatio(
            name_ar="نسبة السيولة الجارية",
            name_en="Current Ratio",
            value=round(current_ratio, 2),
            unit="x",
            benchmark=b["current_ratio"],
            score=s,
            interpretation=f"{'ممتازة — الشركة قادرة على تغطية التزاماتها' if current_ratio >= 1.5 else 'قريبة من الحد الأدنى — يُنصح بتحسين رأس المال العامل'}",
            status=self._status(s)
        ))

        # 7. نسبة السيولة السريعة
        quick_assets = d.current_assets - d.inventory
        quick_ratio = (quick_assets / d.current_liabilities) if d.current_liabilities else 0
        s = self._score(quick_ratio, b["quick_ratio"])
        ratios.append(FinancialRatio(
            name_ar="نسبة السيولة السريعة",
            name_en="Quick Ratio",
            value=round(quick_ratio, 2),
            unit="x",
            benchmark=b["quick_ratio"],
            score=s,
            interpretation="السيولة بعد استبعاد المخزون — مقياس أكثر دقة",
            status=self._status(s)
        ))

        # 8. نسبة النقدية
        cash_ratio = (d.cash / d.current_liabilities) if d.current_liabilities else 0
        s = self._score(cash_ratio, 0.3)
        ratios.append(FinancialRatio(
            name_ar="نسبة النقدية",
            name_en="Cash Ratio",
            value=round(cash_ratio, 2),
            unit="x",
            benchmark=0.3,
            score=s,
            interpretation=f"النقد المتاح فوراً لتغطية الالتزامات",
            status=self._status(s)
        ))

        # ─── نسب الرفع المالي ─────────────────────────────────────────────────

        # 9. نسبة الدين إلى حقوق الملكية
        debt_to_equity = (d.total_liabilities / d.equity) if d.equity else 0
        s = self._score(debt_to_equity, b["debt_to_equity"], higher_is_better=False)
        ratios.append(FinancialRatio(
            name_ar="نسبة الدين إلى حقوق الملكية",
            name_en="Debt to Equity",
            value=round(debt_to_equity, 2),
            unit="x",
            benchmark=b["debt_to_equity"],
            score=s,
            interpretation=f"{'هيكل مالي محافظ' if debt_to_equity < 0.5 else 'اعتماد معتدل على الديون' if debt_to_equity < 1.5 else 'رفع مالي مرتفع'}",
            status=self._status(s)
        ))

        # 10. نسبة الدين إلى الأصول
        debt_to_assets = (d.total_liabilities / d.total_assets) if d.total_assets else 0
        s = self._score(debt_to_assets, 0.5, higher_is_better=False)
        ratios.append(FinancialRatio(
            name_ar="نسبة الدين إلى الأصول",
            name_en="Debt to Assets",
            value=round(debt_to_assets, 2),
            unit="x",
            benchmark=0.5,
            score=s,
            interpretation=f"{debt_to_assets*100:.0f}% من الأصول ممولة بالديون",
            status=self._status(s)
        ))

        # 11. تغطية الفوائد
        interest_coverage = (d.ebit / d.interest_expense) if d.interest_expense else 10
        s = self._score(interest_coverage, b["interest_coverage"])
        ratios.append(FinancialRatio(
            name_ar="نسبة تغطية الفوائد",
            name_en="Interest Coverage",
            value=round(interest_coverage, 2),
            unit="x",
            benchmark=b["interest_coverage"],
            score=s,
            interpretation=f"الأرباح تغطي الفوائد {interest_coverage:.1f}x — {'مريح' if interest_coverage >= 3 else 'يحتاج انتباه'}",
            status=self._status(s)
        ))

        # ─── نسب الكفاءة ──────────────────────────────────────────────────────

        # 12. معدل دوران الأصول
        asset_turnover = (d.revenue / d.total_assets) if d.total_assets else 0
        s = self._score(asset_turnover, b["asset_turnover"])
        ratios.append(FinancialRatio(
            name_ar="معدل دوران الأصول",
            name_en="Asset Turnover",
            value=round(asset_turnover, 2),
            unit="x",
            benchmark=b["asset_turnover"],
            score=s,
            interpretation=f"كل ريال في الأصول يولّد {asset_turnover:.2f} ريال مبيعات",
            status=self._status(s)
        ))

        # 13. أيام تحصيل الديون
        receivable_days = (d.accounts_receivable / d.revenue * self.d) if d.revenue else 0
        s = self._score(receivable_days, b["receivable_days"], higher_is_better=False)
        ratios.append(FinancialRatio(
            name_ar="دورة تحصيل الديون",
            name_en="Days Sales Outstanding",
            value=round(receivable_days, 1),
            unit="day",
            benchmark=b["receivable_days"],
            score=s,
            interpretation=f"متوسط {receivable_days:.0f} يوماً لتحصيل الديون — {'جيد' if receivable_days <= b['receivable_days'] else 'يحتاج تسريع'}",
            status=self._status(s)
        ))

        # 14. دورة المخزون
        if d.cost_of_goods_sold and d.inventory:
            inventory_days = (d.inventory / d.cost_of_goods_sold * self.d)
        else:
            inventory_days = 0
        s = self._score(inventory_days, b["inventory_days"], higher_is_better=False) if inventory_days else 50
        ratios.append(FinancialRatio(
            name_ar="دورة المخزون",
            name_en="Inventory Days",
            value=round(inventory_days, 1),
            unit="day",
            benchmark=b["inventory_days"],
            score=s,
            interpretation=f"المخزون يدور كل {inventory_days:.0f} يوماً",
            status=self._status(s)
        ))

        # 15. معدل نمو الإيرادات
        if d.previous_revenue:
            revenue_growth = ((d.revenue - d.previous_revenue) / d.previous_revenue * 100)
        else:
            revenue_growth = 0
        s = self._score(revenue_growth, 10)  # معيار 10% نمو سنوي
        ratios.append(FinancialRatio(
            name_ar="معدل نمو الإيرادات",
            name_en="Revenue Growth Rate",
            value=round(revenue_growth, 2),
            unit="%",
            benchmark=10.0,
            score=s,
            interpretation=f"نمو {'قوي' if revenue_growth >= 15 else 'معتدل' if revenue_growth >= 5 else 'ضعيف'} مقارنة بالسنة الماضية",
            status=self._status(s)
        ))

        # 16. رأس المال العامل
        working_capital = d.current_assets - d.current_liabilities
        wc_to_assets = (working_capital / d.total_assets * 100) if d.total_assets else 0
        s = self._score(wc_to_assets, 15)
        ratios.append(FinancialRatio(
            name_ar="رأس المال العامل إلى الأصول",
            name_en="Working Capital to Assets",
            value=round(wc_to_assets, 2),
            unit="%",
            benchmark=15.0,
            score=s,
            interpretation=f"رأس المال العامل: {working_capital:,.0f} ريال",
            status=self._status(s)
        ))

        return ratios

    def calculate_readiness_score(self, ratios: list) -> tuple:
        """يحسب درجة الجاهزية الاستثمارية من 100"""

        # أوزان كل فئة
        weights = {
            "profitability": 0.30,   # الربحية
            "liquidity": 0.20,       # السيولة
            "leverage": 0.20,        # الرفع المالي
            "efficiency": 0.15,      # الكفاءة
            "growth": 0.15,          # النمو
        }

        # تصنيف النسب لفئاتها
        profitability_indices = [0, 1, 2, 3, 4]     # هامش الربح، ROE، ROA
        liquidity_indices = [5, 6, 7]               # السيولة
        leverage_indices = [8, 9, 10]               # الديون
        efficiency_indices = [11, 12, 13]           # الكفاءة
        growth_indices = [14, 15]                   # النمو

        def avg_score(indices):
            scores = [ratios[i].score for i in indices if i < len(ratios)]
            return sum(scores) / len(scores) if scores else 50

        total = (
            avg_score(profitability_indices) * weights["profitability"] +
            avg_score(liquidity_indices) * weights["liquidity"] +
            avg_score(leverage_indices) * weights["leverage"] +
            avg_score(efficiency_indices) * weights["efficiency"] +
            avg_score(growth_indices) * weights["growth"]
        )

        score = round(total, 1)

        if score >= 80:
            label = "جاهز للتمويل Series A ✓"
        elif score >= 65:
            label = "جاهز للتمويل الأولي"
        elif score >= 50:
            label = "يحتاج تحسينات قبل التمويل"
        elif score >= 35:
            label = "بحاجة إلى إعادة هيكلة"
        else:
            label = "غير جاهز للتمويل حالياً"

        return score, label

    def generate_insights(self, ratios: list) -> list:
        """يولّد توصيات تلقائية بدون AI"""
        insights = []

        for r in ratios:
            if r.status == "good" and r.score >= 80:
                insights.append({
                    "type": "strength",
                    "title": f"نقطة قوة — {r.name_ar}",
                    "text": r.interpretation,
                    "icon": "trending_up"
                })
            elif r.status == "danger":
                insights.append({
                    "type": "warning",
                    "title": f"يحتاج انتباه — {r.name_ar}",
                    "text": r.interpretation,
                    "icon": "warning"
                })

        # أضف فرصة نمو إذا كانت الإيرادات تنمو
        revenue_growth = next((r for r in ratios if r.name_en == "Revenue Growth Rate"), None)
        if revenue_growth and revenue_growth.value > 10:
            insights.append({
                "type": "opportunity",
                "title": "فرصة — نمو إيرادات قوي",
                "text": f"بمعدل نمو {revenue_growth.value:.1f}%، التوقعات إيجابية للسنوات الثلاث القادمة",
                "icon": "auto_graph"
            })

        return insights[:5]  # أقصى 5 توصيات

    def forecast(self) -> dict:
        """توقعات بسيطة للسنوات 3-5 القادمة"""
        d = self.data
        if not d.revenue:
            return {}

        # معدل نمو مقترح بناءً على الأداء الحالي
        growth_rate = 0.15  # افتراضي 15%
        if d.previous_revenue:
            actual_growth = (d.revenue - d.previous_revenue) / d.previous_revenue
            growth_rate = min(max(actual_growth, 0.05), 0.35)  # بين 5% و35%

        years = [2025, 2026, 2027, 2028, 2029]
        optimistic = []
        base = []
        conservative = []

        rev = d.revenue
        for i, year in enumerate(years):
            rev = rev * (1 + growth_rate)
            optimistic.append({"year": year, "revenue": round(rev * 1.2)})
            base.append({"year": year, "revenue": round(rev)})
            conservative.append({"year": year, "revenue": round(rev * 0.8)})

        return {
            "optimistic": optimistic,
            "base": base,
            "conservative": conservative,
            "growth_rate_used": round(growth_rate * 100, 1)
        }

    def analyze(self) -> AnalysisResult:
        """التحليل الكامل — النقطة الرئيسية"""
        ratios = self.calculate_all()
        score, label = self.calculate_readiness_score(ratios)
        insights = self.generate_insights(ratios)
        forecast = self.forecast()

        return AnalysisResult(
            company_name=self.data.industry,
            period="2024",
            ratios=[asdict(r) for r in ratios],
            readiness_score=score,
            readiness_label=label,
            ai_insights=insights,
            forecast=forecast,
        )


# ─── EXCEL READER ─────────────────────────────────────────────────────────────

def read_from_excel(filepath: str) -> FinancialData:
    """
    يقرأ البيانات المالية من Excel
    
    التنسيق المتوقع: Sheet اسمه "Financial Data"
    مع عمودين: "Item" و "Value"
    
    مثال:
        Item                | Value
        Revenue             | 2400000
        Cost of Goods Sold  | 1380000
        ...
    """
    try:
        import pandas as pd
        
        df = pd.read_excel(filepath, sheet_name=0)
        
        # تنظيف أسماء الأعمدة
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # البحث عن عمود القيمة
        value_col = None
        for col in df.columns:
            if 'value' in col or 'قيمة' in col or col == 'amount':
                value_col = col
                break
        
        if value_col is None and len(df.columns) >= 2:
            value_col = df.columns[1]
        
        # بناء قاموس البيانات
        item_col = df.columns[0]
        data_dict = {}
        for _, row in df.iterrows():
            key = str(row[item_col]).strip().lower() if row[item_col] else ""
            try:
                val = float(str(row[value_col]).replace(',', '').replace(' ', ''))
                data_dict[key] = val
            except:
                pass
        
        # تعيين القيم للحقول
        def get(keys, default=0):
            for k in keys:
                if k in data_dict:
                    return data_dict[k]
            return default
        
        revenue = get(['revenue', 'الإيرادات', 'sales', 'total revenue', 'المبيعات'])
        cogs = get(['cost of goods sold', 'cogs', 'تكلفة المبيعات', 'cost of sales'])
        gross_profit = get(['gross profit', 'الربح الإجمالي']) or (revenue - cogs)
        
        return FinancialData(
            revenue=revenue,
            cost_of_goods_sold=cogs,
            gross_profit=gross_profit,
            operating_expenses=get(['operating expenses', 'المصاريف التشغيلية', 'opex']),
            ebitda=get(['ebitda', 'أرباح قبل الفوائد']),
            ebit=get(['ebit', 'operating profit', 'الربح التشغيلي']),
            interest_expense=get(['interest expense', 'مصاريف الفوائد', 'finance cost']),
            net_profit=get(['net profit', 'net income', 'صافي الربح', 'bottom line']),
            current_assets=get(['current assets', 'الأصول المتداولة']),
            cash=get(['cash', 'النقدية', 'cash and equivalents']),
            accounts_receivable=get(['accounts receivable', 'حسابات القبض', 'receivables']),
            inventory=get(['inventory', 'المخزون', 'stock']),
            total_assets=get(['total assets', 'إجمالي الأصول']),
            current_liabilities=get(['current liabilities', 'الالتزامات المتداولة']),
            total_liabilities=get(['total liabilities', 'إجمالي الالتزامات']),
            equity=get(['equity', 'shareholders equity', 'حقوق الملكية', 'net assets']),
            long_term_debt=get(['long term debt', 'الديون طويلة الأجل']),
            previous_revenue=get(['previous revenue', 'إيرادات السنة الماضية', 'prior year revenue']),
        )
    
    except ImportError:
        print("خطأ: تحتاج تثبيت pandas: pip install pandas openpyxl")
        raise


# ─── SAMPLE DATA (للاختبار بدون Excel) ───────────────────────────────────────

def get_sample_data() -> FinancialData:
    """بيانات تجريبية لشركة صناعية سعودية متوسطة"""
    return FinancialData(
        revenue=2_400_000,
        cost_of_goods_sold=1_380_000,
        gross_profit=1_020_000,
        operating_expenses=660_000,
        ebitda=540_000,
        ebit=360_000,
        interest_expense=86_000,
        net_profit=340_800,
        current_assets=890_000,
        cash=220_000,
        accounts_receivable=280_000,
        inventory=390_000,
        total_assets=2_720_000,
        current_liabilities=490_000,
        total_liabilities=980_000,
        equity=1_740_000,
        long_term_debt=490_000,
        previous_revenue=2_025_000,
        industry="general",
    )


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='APEX Financial Analyzer')
    parser.add_argument('--file', type=str, help='مسار ملف Excel أو PDF')
    parser.add_argument('--sample', action='store_true', help='استخدام بيانات تجريبية')
    parser.add_argument('--output', type=str, default='apex_result.json', help='ملف النتائج')
    args = parser.parse_args()

    print("=" * 60)
    print("APEX Financial Analyzer — أبيكس للتحليل المالي")
    print("=" * 60)

    if args.file:
        print(f"جاري قراءة: {args.file}")
        data = read_from_excel(args.file)
    else:
        print("جاري تحليل البيانات التجريبية...")
        data = get_sample_data()

    analyzer = ApexAnalyzer(data)
    result = analyzer.analyze()

    print(f"\n✓ تم التحليل — {len(result.ratios)} نسبة مالية محسوبة")
    print(f"✓ درجة الجاهزية: {result.readiness_score}/100 — {result.readiness_label}")
    print(f"\nالنسب المالية:")
    print("-" * 50)

    for r in result.ratios:
        status_icon = "✓" if r['status'] == 'good' else "⚠" if r['status'] == 'warning' else "✗"
        print(f"  {status_icon} {r['name_ar']}: {r['value']}{r['unit']} (درجة: {r['score']}/100)")

    # حفظ النتائج
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(result.__dict__, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ تم حفظ النتائج في: {args.output}")
    print("=" * 60)

    return result


if __name__ == "__main__":
    main()

def read_trial_balance(filepath: str) -> FinancialData:
    """يقرأ ميزان المراجعة ويعتمد على التبويب في العمود B"""
    from openpyxl import load_workbook
    import warnings
    warnings.filterwarnings('ignore')

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row[1] or not row[2]:
            continue
        def to_float(v):
            try:
                if isinstance(v, (int, float)):
                    return float(v)
                if v is None:
                    return 0.0
                s = str(v).strip().replace(',', '')
                return float(s) if s else 0.0
            except:
                return 0.0
        d = to_float(row[3])
        e = to_float(row[4])
        f = to_float(row[5])
        g = to_float(row[6])
        net = (d + f) - (e + g)
        rows.append({'tab': str(row[1]).strip(), 'name': str(row[2]).strip(), 'net': net, 'fin_d': d+f, 'fin_c': e+g})

    def sm(keywords):
        return sum(abs(r['net']) for r in rows if any(k in r['tab'] for k in keywords))

    def snet(keywords):
        return sum(r['net'] for r in rows if any(k in r['tab'] for k in keywords))

    # قائمة الدخل
    revenue = sm(['إيرادات - مبيعات محلية'])
    other_income = sm(['إيرادات - إيرادات خدمات'])
    sales_discount = sm(['إيرادات - خصم مبيعات'])
    sales_returns = sm(['إيرادات - مردودات ومسموحات مبيعات'])
    net_revenue = revenue + other_income + sales_discount - sales_returns

    cogs = sm(['تكلفة مبيعات - تكلفة بضاعة مباعة'])
    gross_profit = net_revenue - cogs

    opex = sm(['مصروفات إدارية', 'مصروفات بيع وتوزيع'])
    ebit = gross_profit - opex

    interest = sm(['تكاليف تمويل'])
    tax = sm(['زكاة وضرائب'])
    net_profit = ebit - interest - tax

    # الميزانية
    cash = sm(['أصول متداولة - نقد وما في حكمه'])
    receivables = sm(['أصول متداولة - ذمم مدينة تجارية', 'أصول متداولة - ذمم مدينة أخرى'])
    inventory = sm(['أصول متداولة - مخزون'])
    prepaid = sm(['أصول متداولة - مصروفات مدفوعة'])
    vat_in = sm(['أصول متداولة - ضريبة قيمة مضافة مدخلات'])
    current_assets = cash + receivables + inventory + prepaid + vat_in

    fixed_gross = sm(['أصول غير متداولة - ممتلكات وآلات ومعدات'])
    depreciation = sm(['أصول غير متداولة - مجمع الإهلاك'])
    fixed_assets = max(fixed_gross - depreciation, 0)
    total_assets = current_assets + fixed_assets

    current_liabilities = sm(['التزامات متداولة'])
    lt_liabilities = sm(['التزامات غير متداولة'])
    total_liabilities = current_liabilities + lt_liabilities

    equity = sm(['حقوق ملكية - رأس المال المدفوع', 'حقوق ملكية - احتياطي نظامي', 'حقوق ملكية - أرباح مبقاة سنوات سابقة'])
    if equity == 0:
        equity = max(total_assets - total_liabilities, 0)

    return FinancialData(
        revenue=net_revenue,
        cost_of_goods_sold=cogs,
        gross_profit=gross_profit,
        operating_expenses=opex,
        ebit=ebit,
        ebitda=ebit * 1.1,
        interest_expense=interest,
        net_profit=net_profit,
        current_assets=current_assets,
        cash=cash,
        inventory=inventory,
        total_assets=total_assets,
        current_liabilities=current_liabilities,
        total_liabilities=total_liabilities,
        equity=equity,
    )

def read_from_excel(filepath: str) -> FinancialData:
    """
    يقرأ البيانات المالية من Excel
    
    التنسيق المتوقع: Sheet اسمه "Financial Data"
    مع عمودين: "Item" و "Value"
    
    مثال:
        Item                | Value
        Revenue             | 2400000
        Cost of Goods Sold  | 1380000
        ...
    """
    try:
        import pandas as pd
        
        df = pd.read_excel(filepath, sheet_name=0)
        
        # تنظيف أسماء الأعمدة
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # البحث عن عمود القيمة
        value_col = None
        for col in df.columns:
            if 'value' in col or 'قيمة' in col or col == 'amount':
                value_col = col
                break
        
        if value_col is None and len(df.columns) >= 2:
            value_col = df.columns[1]
        
        # بناء قاموس البيانات
        item_col = df.columns[0]
        data_dict = {}
        for _, row in df.iterrows():
            key = str(row[item_col]).strip().lower() if row[item_col] else ""
            try:
                val = float(str(row[value_col]).replace(',', '').replace(' ', ''))
                data_dict[key] = val
            except:
                pass
        
        # تعيين القيم للحقول
        def get(keys, default=0):
            for k in keys:
                if k in data_dict:
                    return data_dict[k]
            return default
        
        revenue = get(['revenue', 'الإيرادات', 'sales', 'total revenue', 'المبيعات'])
        cogs = get(['cost of goods sold', 'cogs', 'تكلفة المبيعات', 'cost of sales'])
        gross_profit = get(['gross profit', 'الربح الإجمالي']) or (revenue - cogs)
        
        return FinancialData(
            revenue=revenue,
            cost_of_goods_sold=cogs,
            gross_profit=gross_profit,
            operating_expenses=get(['operating expenses', 'المصاريف التشغيلية', 'opex']),
            ebitda=get(['ebitda', 'أرباح قبل الفوائد']),
            ebit=get(['ebit', 'operating profit', 'الربح التشغيلي']),
            interest_expense=get(['interest expense', 'مصاريف الفوائد', 'finance cost']),
            net_profit=get(['net profit', 'net income', 'صافي الربح', 'bottom line']),
            current_assets=get(['current assets', 'الأصول المتداولة']),
            cash=get(['cash', 'النقدية', 'cash and equivalents']),
            accounts_receivable=get(['accounts receivable', 'حسابات القبض', 'receivables']),
            inventory=get(['inventory', 'المخزون', 'stock']),
            total_assets=get(['total assets', 'إجمالي الأصول']),
            current_liabilities=get(['current liabilities', 'الالتزامات المتداولة']),
            total_liabilities=get(['total liabilities', 'إجمالي الالتزامات']),
            equity=get(['equity', 'shareholders equity', 'حقوق الملكية', 'net assets']),
            long_term_debt=get(['long term debt', 'الديون طويلة الأجل']),
            previous_revenue=get(['previous revenue', 'إيرادات السنة الماضية', 'prior year revenue']),
        )
    
    except ImportError:
        print("خطأ: تحتاج تثبيت pandas: pip install pandas openpyxl")
        raise


# ─── SAMPLE DATA (للاختبار بدون Excel) ───────────────────────────────────────

def get_sample_data() -> FinancialData:
    """بيانات تجريبية لشركة صناعية سعودية متوسطة"""
    return FinancialData(
        revenue=2_400_000,
        cost_of_goods_sold=1_380_000,
        gross_profit=1_020_000,
        operating_expenses=660_000,
        ebitda=540_000,
        ebit=360_000,
        interest_expense=86_000,
        net_profit=340_800,
        current_assets=890_000,
        cash=220_000,
        accounts_receivable=280_000,
        inventory=390_000,
        total_assets=2_720_000,
        current_liabilities=490_000,
        total_liabilities=980_000,
        equity=1_740_000,
        long_term_debt=490_000,
        previous_revenue=2_025_000,
        industry="general",
    )


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='APEX Financial Analyzer')
    parser.add_argument('--file', type=str, help='مسار ملف Excel أو PDF')
    parser.add_argument('--sample', action='store_true', help='استخدام بيانات تجريبية')
    parser.add_argument('--output', type=str, default='apex_result.json', help='ملف النتائج')
    args = parser.parse_args()

    print("=" * 60)
    print("APEX Financial Analyzer — أبيكس للتحليل المالي")
    print("=" * 60)

    if args.file:
        print(f"جاري قراءة: {args.file}")
        data = read_from_excel(args.file)
    else:
        print("جاري تحليل البيانات التجريبية...")
        data = get_sample_data()

    analyzer = ApexAnalyzer(data)
    result = analyzer.analyze()

    print(f"\n✓ تم التحليل — {len(result.ratios)} نسبة مالية محسوبة")
    print(f"✓ درجة الجاهزية: {result.readiness_score}/100 — {result.readiness_label}")
    print(f"\nالنسب المالية:")
    print("-" * 50)

    for r in result.ratios:
        status_icon = "✓" if r['status'] == 'good' else "⚠" if r['status'] == 'warning' else "✗"
        print(f"  {status_icon} {r['name_ar']}: {r['value']}{r['unit']} (درجة: {r['score']}/100)")

    # حفظ النتائج
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(result.__dict__, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ تم حفظ النتائج في: {args.output}")
    print("=" * 60)

    return result


if __name__ == "__main__":
    main()

def read_trial_balance(filepath: str) -> FinancialData:
    """يقرأ ميزان المراجعة ويعتمد على عمود الرصيد النهائي"""
    import pandas as pd

    # الملف عنده header في صفين (5 و 6) — نقرأ من الصف 5
    df = pd.read_excel(filepath, sheet_name=0, header=[4, 5])

    # تنظيف أسماء الأعمدة
    df.columns = [str(c).strip() for c in df.columns]

    # البحث عن عمود الرصيد بعد التعديلات (مدين ودائن)
    adj_debit_col = None
    adj_credit_col = None
    code_col = None
    account_col = None

    for col in df.columns:
        col_clean = col.strip()
        # عمود الكود
        if any(k in col_clean for k in ['كود', 'رمز', 'code', 'Code', 'رقم الحساب']):
            code_col = col
        # عمود اسم الحساب
        if any(k in col_clean for k in ['اسم', 'الحساب', 'account', 'Account', 'البيان']):
            account_col = col
        # عمود الرصيد بعد التعديلات مدين
        if ('بعد' in col_clean or 'تعديل' in col_clean) and ('مدين' in col_clean or 'debit' in col_clean.lower()):
            adj_debit_col = col
        # عمود الرصيد بعد التعديلات دائن
        if ('بعد' in col_clean or 'تعديل' in col_clean) and ('دائن' in col_clean or 'credit' in col_clean.lower()):
            adj_credit_col = col

    # fallback: لو ما لقى عمود التعديلات يستخدم آخر عمودين
    if adj_debit_col is None or adj_credit_col is None:
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if len(numeric_cols) >= 2:
            adj_debit_col = numeric_cols[-2]
            adj_credit_col = numeric_cols[-1]

    # fallback للكود
    if code_col is None:
        code_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]

    def g(code_val):
        """جلب قيمة حساب معين بالكود"""
        try:
            row = df[df[code_col].astype(str).str.strip() == str(code_val)]
            if not row.empty:
                d = float(row.iloc[0][adj_debit_col]) if adj_debit_col and pd.notna(row.iloc[0][adj_debit_col]) else 0
                c = float(row.iloc[0][adj_credit_col]) if adj_credit_col and pd.notna(row.iloc[0][adj_credit_col]) else 0
                return d, c
        except Exception:
            pass
        return 0, 0

    def g_sum(code_prefix):
        """جمع كل الحسابات التي تبدأ بكود معين"""
        try:
            mask = df[code_col].astype(str).str.strip().str.startswith(str(code_prefix))
            rows = df[mask]
            if not rows.empty:
                d = rows[adj_debit_col].apply(lambda x: float(x) if pd.notna(x) else 0).sum() if adj_debit_col else 0
                c = rows[adj_credit_col].apply(lambda x: float(x) if pd.notna(x) else 0).sum() if adj_credit_col else 0
                return d, c
        except Exception:
            pass
        return 0, 0

    revenue = g('42')[1]
    cogs = g('32')[0]
    gross_profit = revenue - cogs
    opex = g('31')[0]
    ebit = gross_profit - opex
    interest = g('310065')[0]
    net_profit = ebit - interest
    total_assets = g('1')[0]
    current_assets = g('122')[0]
    cash = g('125')[0]
    inventory = g('128')[0]
    total_liabilities = g('2')[1]
    current_liabilities = g('21')[1]
    equity = g('3')[0]

    return FinancialData(
        revenue=revenue,
        cost_of_goods_sold=cogs,
        gross_profit=gross_profit,
        operating_expenses=opex,
        ebit=ebit,
        ebitda=ebit * 1.1,
        interest_expense=interest,
        net_profit=net_profit,
        current_assets=current_assets,
        cash=cash,
        inventory=inventory,
        total_assets=total_assets,
        current_liabilities=current_liabilities,
        total_liabilities=total_liabilities,
        equity=equity,
    )





