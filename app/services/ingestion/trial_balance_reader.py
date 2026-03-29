"""
APEX Ingestion Service — قراءة الملفات وتطبيع البيانات
═══════════════════════════════════════════════════════════

يدعم 3 أنواع من النماذج:
  1. apex_v1: النموذج القديم (12 عمود) — B=tab, C=name, D-L=numbers
  2. apex_v2: النموذج الجديد (10 عمود) — A=code, B=main_tab, C=sub_tab, D=name, E-J=numbers
  3. generic: أي ملف Excel آخر → يُعامل كـ v1
"""

from openpyxl import load_workbook


class TrialBalanceReader:

    def read(self, filepath: str) -> dict:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        warnings = []

        format_type = self._detect_format(ws)
        meta = self._read_meta(ws)

        if format_type == "apex_v2":
            rows = self._read_v2_format(ws, warnings)
        else:
            rows = self._read_v1_format(ws, warnings)

        wb.close()
        rows = self._filter_summary_rows(rows)

        return {
            "rows": rows,
            "meta": meta,
            "format": format_type,
            "row_count": len(rows),
            "warnings": warnings,
        }

    def _detect_format(self, ws) -> str:
        """
        apex_v1: Row 5-6 headers with 'تبويب الحساب' — 12 columns — data row 7
        apex_v2: Row 7-8 headers with 'التبويب الرئيسي' — 10 columns — data row 9
        """
        for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
            if not row:
                continue
            for cell in row:
                if cell and isinstance(cell, str):
                    if "تبويب الحساب" in cell or "Account Classific" in cell:
                        return "apex_v1"

        for row in ws.iter_rows(min_row=7, max_row=8, values_only=True):
            if not row:
                continue
            for cell in row:
                if cell and isinstance(cell, str):
                    if "التبويب الرئيسي" in cell or "Main Category" in cell:
                        return "apex_v2"

        # Fallback: check first data row
        for row in ws.iter_rows(min_row=7, max_row=8, values_only=True):
            if row and len(row) >= 3:
                b_val = row[1]
                if b_val and isinstance(b_val, str):
                    if any(kw in str(b_val) for kw in ["أصول", "التزامات", "حقوق", "إيرادات", "مصروفات", "تكلفة"]):
                        return "apex_v1"
            break

        return "apex_v1"

    def _read_meta(self, ws) -> dict:
        meta = {"company_name": "", "period": "", "currency": "SAR"}
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            if not row:
                continue
            for cell in row:
                if cell and isinstance(cell, str):
                    text = cell.strip()
                    if any(kw in text for kw in ["شركة", "مؤسسة", "مجموعة", "Company"]):
                        meta["company_name"] = text
                    if any(kw in text for kw in ["2024", "2025", "2026", "Dec", "السنة", "الفترة"]):
                        if not meta["period"]:
                            meta["period"] = text
        return meta

    def _read_v1_format(self, ws, warnings: list) -> list:
        """
        APEX v1 (12 columns): B=tab, C=name, D=open_dr, E=open_cr,
        F=mov_dr, G=mov_cr, H=close_dr, I=close_cr,
        J=bal_before, K=adj, L=bal_after. Data from row 7.
        """
        rows = []
        for row_data in ws.iter_rows(min_row=7, values_only=True):
            if not row_data:
                continue

            tab = row_data[1] if len(row_data) > 1 else None
            name = row_data[2] if len(row_data) > 2 else None

            if not tab or not name:
                continue
            if not isinstance(name, str) or not name.strip():
                continue

            tab_clean = str(tab).strip()
            name_clean = str(name).strip()

            open_d = self._to_float(row_data[3]) if len(row_data) > 3 else 0.0
            open_c = self._to_float(row_data[4]) if len(row_data) > 4 else 0.0
            mov_d = self._to_float(row_data[5]) if len(row_data) > 5 else 0.0
            mov_c = self._to_float(row_data[6]) if len(row_data) > 6 else 0.0
            close_d = self._to_float(row_data[7]) if len(row_data) > 7 else 0.0
            close_c = self._to_float(row_data[8]) if len(row_data) > 8 else 0.0

            # Column L (index 11) = adjusted balance — primary source
            adj = self._to_float(row_data[11]) if len(row_data) > 11 else None

            if adj is not None and adj != 0:
                net = adj
            elif close_d != 0 or close_c != 0:
                net = close_d - close_c
            else:
                net = (open_d - open_c) + (mov_d - mov_c)

            rows.append({
                "code": str(row_data[0]).strip() if row_data[0] else "",
                "tab": tab_clean,
                "sub_tab": "",
                "name": name_clean,
                "open_debit": open_d,
                "open_credit": open_c,
                "movement_debit": mov_d,
                "movement_credit": mov_c,
                "close_debit": close_d,
                "close_credit": close_c,
                "net_balance": net,
            })

        if not rows:
            warnings.append("لم يتم العثور على بيانات في الملف")
        return rows

    def _read_v2_format(self, ws, warnings: list) -> list:
        """
        APEX v2 (10 columns): A=code, B=main_tab, C=sub_tab, D=name,
        E=open_dr, F=open_cr, G=mov_dr, H=mov_cr,
        I=close_dr, J=close_cr. Data from row 9.
        """
        rows = []
        for row_data in ws.iter_rows(min_row=9, values_only=True):
            if not row_data or len(row_data) < 8:
                continue

            name = row_data[3]
            if not name or not str(name).strip():
                continue

            tab_raw = str(row_data[1]).strip() if row_data[1] else ""
            sub = str(row_data[2]).strip() if row_data[2] else ""
            code = str(row_data[0]).strip() if row_data[0] else ""

            open_d = self._to_float(row_data[4])
            open_c = self._to_float(row_data[5])
            mov_d = self._to_float(row_data[6])
            mov_c = self._to_float(row_data[7])
            net = (open_d - open_c) + (mov_d - mov_c)

            rows.append({
                "code": code,
                "tab": tab_raw,
                "sub_tab": sub,
                "name": str(name).strip(),
                "open_debit": open_d,
                "open_credit": open_c,
                "movement_debit": mov_d,
                "movement_credit": mov_c,
                "close_debit": max(net, 0),
                "close_credit": abs(min(net, 0)),
                "net_balance": net,
            })

        if not rows:
            warnings.append("لم يتم العثور على بيانات في الملف")
        return rows

    def _filter_summary_rows(self, rows: list) -> list:
        """Remove total/summary rows and category header rows."""
        skip_name_keywords = [
            "فحص التوازن", "balance check", "✓", "✗",
            "متوازن", "غير متوازن",
        ]

        # Generic top-level tabs that are category headers, not real accounts
        summary_tabs = [
            "الأصول", "الالتزامات", "حقوق الملكية", "الإيرادات",
            "تكلفة المبيعات", "المصروفات التشغيلية", "تكاليف التمويل",
            "الزكاة والضرائب", "المصروفات", "الإيرادات الأخرى",
        ]

        filtered = []
        for row in rows:
            name = row.get("name", "").strip()
            tab = row.get("tab", "").strip()

            if not name:
                continue

            # Skip rows with name keywords
            if any(kw in name.lower() for kw in skip_name_keywords):
                continue

            # Skip category header rows: tab is a generic label without " - "
            # AND name contains "|" or ":" (looks like a category summary)
            if tab in summary_tabs:
                continue

            # Skip rows where net_balance is exactly 0 and name contains numbering pattern
            if "|" in name and ":" in name:
                continue

            filtered.append(row)
        return filtered

    @staticmethod
    def _to_float(v) -> float:
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.replace(",", "").strip())
            except (ValueError, AttributeError):
                pass
        return 0.0
