"""
APEX Sprint 1 — COA File Reader
═══════════════════════════════════════════════════════════════
Reads CSV/XLSX/XLS files, detects sheets, headers, columns.
Per Sprint 1 Build Spec §17.1.
"""

import os
import csv
from typing import Optional, List, Dict, Any, Iterable

# ── Column Aliases (Arabic + English) per Sprint 1 Spec §8.3 ──

COLUMN_ALIASES = {
    "account_code": [
        "account_code",
        "code",
        "acc_code",
        "account number",
        "account_no",
        "رقم الحساب",
        "كود الحساب",
        "رقم_الحساب",
        "الرقم",
        "رمز الحساب",
    ],
    "account_name": [
        "account_name",
        "name",
        "account title",
        "description",
        "account_title",
        "اسم الحساب",
        "اسم_الحساب",
        "البيان",
        "الوصف",
        "مسمى الحساب",
        "اسم",
    ],
    "parent_code": [
        "parent_code",
        "parent",
        "main_code",
        "parent account code",
        "parent_account",
        "كود الأب",
        "رقم الحساب الأب",
        "الحساب الرئيسي",
        "رقم_الأب",
    ],
    "parent_name": [
        "parent_name",
        "parent account name",
        "parent_account_name",
        "اسم الأب",
        "اسم الحساب الأب",
        "اسم_الأب",
    ],
    "level": [
        "level",
        "account_level",
        "lvl",
        "depth",
        "المستوى",
        "مستوى",
        "العمق",
    ],
    "account_type": [
        "account_type",
        "type",
        "category",
        "classification",
        "نوع الحساب",
        "النوع",
        "التصنيف",
        "الفئة",
    ],
    "normal_balance": [
        "normal_balance",
        "balance_type",
        "debit_credit",
        "nature",
        "طبيعة الرصيد",
        "الطبيعة",
        "مدين_دائن",
        "طبيعة",
    ],
    "active_flag": [
        "active",
        "is_active",
        "enabled",
        "active_flag",
        "status",
        "مفعل",
        "نشط",
        "الحالة",
        "فعال",
    ],
    "notes": [
        "notes",
        "remarks",
        "comments",
        "memo",
        "ملاحظات",
        "ملاحظه",
        "تعليقات",
        "بيان إضافي",
    ],
}


def _match_column(raw_col: str) -> Optional[str]:
    """Match a raw column header to a standard field."""
    raw = raw_col.strip().lower()
    # Remove diacritics for matching
    import re

    raw_clean = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670]", "", raw)

    for standard_field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if raw_clean == alias.lower() or raw == alias.lower():
                return standard_field
    return None


def list_sheets(file_path: str) -> List[str]:
    """List sheets in Excel file. Returns ['Sheet1'] for CSV."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        pass

    try:
        import xlrd

        wb = xlrd.open_workbook(file_path)
        return wb.sheet_names()
    except Exception:
        return []


def detect_header_and_columns(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row_index: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Read first rows, detect header, suggest column mapping.
    Returns: {
        detected_columns: [...],
        suggested_mapping: {standard -> raw},
        sample_rows: [...],
        header_row_index: int,
        warnings: [...]
    }
    """
    ext = os.path.splitext(file_path)[1].lower()
    rows = _read_first_rows(file_path, ext, sheet_name, max_rows=15)

    if not rows:
        return {
            "detected_columns": [],
            "suggested_mapping": {},
            "sample_rows": [],
            "header_row_index": 0,
            "warnings": ["empty_file"],
        }

    # Auto-detect header row (try first 5 rows)
    best_row_idx = header_row_index if header_row_index is not None else 0
    best_score = 0

    if header_row_index is None:
        for i in range(min(5, len(rows))):
            row = rows[i]
            score = sum(1 for cell in row if _match_column(str(cell or "")) is not None)
            if score > best_score:
                best_score = score
                best_row_idx = i

    headers = [str(cell or f"col_{j}").strip() for j, cell in enumerate(rows[best_row_idx])]

    # Build suggested mapping
    suggested = {}
    used_raw = set()
    for raw_col in headers:
        match = _match_column(raw_col)
        if match and match not in suggested and raw_col not in used_raw:
            suggested[match] = raw_col
            used_raw.add(raw_col)

    # Fill unmapped standard fields with None
    for std_field in COLUMN_ALIASES:
        if std_field not in suggested:
            suggested[std_field] = None

    # Sample rows (after header)
    sample = []
    for row in rows[best_row_idx + 1 : best_row_idx + 6]:
        sample.append(
            {headers[j]: (cell if cell is not None else "") for j, cell in enumerate(row) if j < len(headers)}
        )

    warnings = []
    if "account_name" not in suggested or suggested["account_name"] is None:
        warnings.append("required_column_missing: account_name not detected")

    return {
        "detected_columns": headers,
        "suggested_mapping": suggested,
        "sample_rows": sample,
        "header_row_index": best_row_idx,
        "warnings": warnings,
    }


def stream_rows(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row_index: int = 0,
) -> Iterable[Dict[str, Any]]:
    """Yield rows as dicts with header keys."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        yield from _stream_csv(file_path, header_row_index)
    elif ext == ".xlsx":
        yield from _stream_xlsx(file_path, sheet_name, header_row_index)
    elif ext == ".xls":
        yield from _stream_xls(file_path, sheet_name, header_row_index)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ── Internal readers ──


def _read_first_rows(file_path, ext, sheet_name=None, max_rows=15):
    """Read first N rows from any supported file type."""
    if ext == ".csv":
        return _read_csv_rows(file_path, max_rows)
    elif ext == ".xlsx":
        return _read_xlsx_rows(file_path, sheet_name, max_rows)
    elif ext == ".xls":
        return _read_xls_rows(file_path, sheet_name, max_rows)
    return []


def _read_csv_rows(path, max_rows):
    rows = []
    encodings = ["utf-8-sig", "utf-8", "cp1256", "latin-1"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    rows.append(row)
            return rows
        except (UnicodeDecodeError, UnicodeError):
            continue
    return rows


def _read_xlsx_rows(path, sheet_name, max_rows):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row))
    wb.close()
    return rows


def _read_xls_rows(path, sheet_name, max_rows):
    import xlrd

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
    rows = []
    for i in range(min(max_rows, ws.nrows)):
        rows.append(ws.row_values(i))
    return rows


def _stream_csv(path, header_row_index):
    encodings = ["utf-8-sig", "utf-8", "cp1256", "latin-1"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                headers = None
                for i, row in enumerate(reader):
                    if i < header_row_index:
                        continue
                    if i == header_row_index:
                        headers = [str(c).strip() for c in row]
                        continue
                    if headers:
                        yield {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
            return
        except (UnicodeDecodeError, UnicodeError):
            continue


def _stream_xlsx(path, sheet_name, header_row_index):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row_index:
            continue
        if i == header_row_index:
            headers = [str(c or f"col_{j}").strip() for j, c in enumerate(row)]
            continue
        if headers:
            yield {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
    wb.close()


def _stream_xls(path, sheet_name, header_row_index):
    import xlrd

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
    headers = None
    for i in range(ws.nrows):
        row = ws.row_values(i)
        if i < header_row_index:
            continue
        if i == header_row_index:
            headers = [str(c).strip() for c in row]
            continue
        if headers:
            yield {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
