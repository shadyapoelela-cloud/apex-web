"""
APEX COA Engine v4.3 — Lexicon Loader (100%)
=============================================
يحمّل المعجم من Excel ويبني فهارس بحث سريعة.
بدون Excel: يعمل من 97 نمط مدمج (AR + EN).

المصدر: APEX_COA_Names_Lexicon_v4_4.xlsx
"""
from __future__ import annotations
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

# ── تطبيع النص العربي (القسم 6.3) ────────────────────────────
_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
_NORM_MAP = str.maketrans({
    "أ":"ا","إ":"ا","آ":"ا","ٱ":"ا",
    "ة":"ه","ى":"ي","ؤ":"و","ئ":"ي","ء":"","ـ":"",
})
_TASHKEEL = re.compile(
    r"[\u064B-\u065F\u0670\u06D6-\u06DC\u06DF-\u06E4\u06E7\u06E8\u06EA-\u06ED]"
)

def normalize_ar(text: str) -> str:
    if not text: return ""
    s = str(text).strip()
    s = _TASHKEEL.sub("", s)
    s = s.translate(_ARABIC_DIGITS)
    s = s.translate(_NORM_MAP)
    return re.sub(r"\s+", " ", s).strip().lower()

def normalize_en(text: str) -> str:
    if not text: return ""
    return re.sub(r"\s+", " ", str(text).strip().lower())

def has_valid_arabic(text: str) -> bool:
    """ملحق ب — تحقق من وجود عربية سليمة."""
    arabic_chars = sum(1 for c in text if "\u0600" <= c <= "\u06FF")
    return arabic_chars > 10


# ── Data Classes ──────────────────────────────────────────────
@dataclass
class ConceptMatch:
    concept_id: str
    confidence: float
    method: str
    matched_on: str = ""

@dataclass
class ConceptEntry:
    concept_id:  str
    code:        str
    name_ar:     str
    name_en:     str
    section:     str
    nature:      str
    level:       str
    ar_variants: Set[str] = field(default_factory=set)
    en_variants: Set[str] = field(default_factory=set)


# ── الأنماط المدمجة (القسم 6.3 — ARABIC_PATTERNS) ────────────
ARABIC_PATTERNS: List[Tuple[re.Pattern, str]] = [
    # نقدية
    (re.compile(r"نقد(?:ية)?(?:\s|$|و|في|لدى)|^النقد$|نقدي$", re.I),         "CASH"),
    (re.compile(r"صندوق.*(نثر|عهد|مصروف)", re.I),                              "PETTY_CASH"),
    (re.compile(r"صندوق|خزين|كاش(?!.*بنك)", re.I),                             "CASH"),
    (re.compile(r"بنك.*(راجح|أهل|رياض|بلاد|إنماء|ساب|فرنس|إسلام|عربي)", re.I),"BANK"),
    (re.compile(r"بنك|مصرف|حساب.?بنك|ودائع.?بنك", re.I),                       "BANK"),
    (re.compile(r"شيك.*تحصيل|أوراق.*تحصيل", re.I),                             "CHECKS_UNDER_COLLECTION"),
    (re.compile(r"نقدية.?مقيد|أموال.?محجوز|ودائع.?مقيد", re.I),               "RESTRICTED_CASH"),
    # مجمعات إهلاك
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(بناء|مبنى|مبان|مباني|إنشاء)", re.I), "ACCUM_DEPR_BUILDINGS"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(آل|معد|ماكين)", re.I),        "ACCUM_DEPR_MACHINERY"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(أثاث|مفروش)", re.I),           "ACCUM_DEPR_FURNITURE"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(سيار|مركب|نقل)", re.I),        "ACCUM_DEPR_VEHICLES"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(حاسب|كمبيوتر|IT)", re.I),     "ACCUM_DEPR_COMPUTERS"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك).*(حق|ROU|استخدام)", re.I),       "ACCUM_DEPR_ROU"),
    (re.compile(r"مجمع.*(إهلاك|اهلاك|استهلاك)", re.I),                         "ACCUM_DEPR_GENERAL"),
    (re.compile(r"مجمع.*(إطفاء|اطفاء)", re.I),                                  "ACCUM_AMORT"),
    # ذمم مدينة
    (re.compile(r"مخصص.*(ائتمان|ديون|معدوم|ECL|مشكوك)", re.I),                 "ECL_PROVISION"),
    (re.compile(r"ذمم.*مدين|مدينون|عملاء|حسابات.*قبض|مستحقات.*عميل", re.I),   "ACC_RECEIVABLE"),
    (re.compile(r"أوراق.?قبض|كمبيالات.?قبض|سندات.?قبض", re.I),                 "NOTES_RECEIVABLE"),
    (re.compile(r"إيرادات.?مستحق|إيرادات.?مستحقة.?قبض", re.I),                 "ACCRUED_REVENUE"),
    (re.compile(r"سلف.?موظف|عهد.?موظف|سلفيات", re.I),                          "EMPLOYEE_RECEIVABLE"),
    # مخزون
    (re.compile(r"مواد.?خام|خامات|مواد.?أولية", re.I),                          "RAW_MATERIALS"),
    (re.compile(r"إنتاج.?(تحت|جار|نصف)|WIP", re.I),                            "WIP"),
    (re.compile(r"بضاع.?(تام|جاهز|مكتمل)|منتجات.?تام", re.I),                  "FINISHED_GOODS"),
    (re.compile(r"مخزون|بضاع(?!.*مباع)|سلع(?!.*تام)", re.I),                    "INVENTORY"),
    # مدفوعات مقدمة
    (re.compile(r"تأمين.?مدفوع.?مقدم|أقساط.?تأمين.?مقدم", re.I),              "PREPAID_INSURANCE"),
    (re.compile(r"إيجار.?مدفوع.?مقدم|أجرة.?مدفوع", re.I),                      "PREPAID_RENT"),
    (re.compile(r"دفع.?مقدم.?(مورد|مشتريات)|سلف.?مورد", re.I),                 "ADVANCE_TO_SUPPLIERS"),
    (re.compile(r"مصروف.?مدفوع.?مقدم|مقدمات(?!.*عميل)", re.I),                 "PREPAID_EXPENSES"),
    # ضرائب
    (re.compile(r"ض.?ق.?م.*(مدخل|مشتريات|input)|VAT.?input", re.I),            "VAT_INPUT"),
    (re.compile(r"ض.?ق.?م.*(مخرج|مبيعات|output)|VAT.?output", re.I),           "VAT_OUTPUT"),
    (re.compile(r"زكاة.*مستحق|مستحق.*زكاة", re.I),                              "ZAKAT_PAYABLE"),
    (re.compile(r"ضريبة.*دخل.*مستحق|income.?tax.?payable", re.I),               "INCOME_TAX_PAYABLE"),
    (re.compile(r"استقطاع.*مورد|ضريبة.*استقطاع", re.I),                         "WITHHOLDING_TAX"),
    # أصول ثابتة
    (re.compile(r"أرض|أراضي|أراضٍ", re.I),                                      "LAND"),
    (re.compile(r"مبان|مبنى|مباني|إنشاءات", re.I),                              "BUILDINGS"),
    (re.compile(r"معدات|آلات|ماكينات", re.I),                                    "MACHINERY_EQUIPMENT"),
    (re.compile(r"أثاث|مفروشات", re.I),                                          "FURNITURE_FIXTURES"),
    (re.compile(r"سيار|مركبات|سيارات|عربات", re.I),                              "VEHICLES"),
    (re.compile(r"حاسب|حواسب|كمبيوتر|أجهزة.*تقنية", re.I),                      "COMPUTERS_IT"),
    (re.compile(r"حق.*استخدام|ROU|IFRS.?16", re.I),                              "ROU_ASSET"),
    (re.compile(r"شهرة.*محل|goodwill", re.I),                                    "GOODWILL"),
    # ذمم دائنة
    (re.compile(r"ذمم.*دائن|دائنون|حسابات.*دفع|مستحق.*مورد", re.I),             "ACC_PAYABLE"),
    (re.compile(r"أوراق.?دفع|كمبيالات.?دفع|سندات.?دفع", re.I),                  "NOTES_PAYABLE"),
    (re.compile(r"مصروف.*مستحق|مستحقات.*دفع(?!.*عميل)", re.I),                  "ACCRUED_EXPENSES"),
    (re.compile(r"سلف.*عملاء|مقدمات.*عملاء|دفعات.*مقدمة.*عميل", re.I),          "CUSTOMER_ADVANCES"),
    (re.compile(r"ضمانات.*نقدية|كفالات.*بنكية", re.I),                           "CASH_BONDS"),
    # IFRS 16
    (re.compile(r"التزام.*إيجار.*متداول|إيجار.*مستحق.*قصير", re.I),             "LEASE_LIABILITY_CURRENT"),
    (re.compile(r"التزام.*إيجار|lease.?liability", re.I),                         "LEASE_LIABILITY_NC"),
    # قروض
    (re.compile(r"قرض.*(بنك|مصرف|طويل|مدة)|تمويل.*طويل", re.I),                 "LONG_TERM_LOAN"),
    (re.compile(r"قرض.*(قصير|جار|متداول)|بنك.*متداول", re.I),                    "SHORT_TERM_LOAN"),
    # حقوق ملكية
    (re.compile(r"رأس.*مال(?!.*ع)", re.I),                                         "PAID_IN_CAPITAL"),
    (re.compile(r"احتياطي.?نظامي|احتياطي.?قانوني", re.I),                        "LEGAL_RESERVE"),
    (re.compile(r"أرباح.?مبقاة|أرباح.?محتجز|retained", re.I),                    "RETAINED_EARNINGS"),
    (re.compile(r"أرباح.?موزع|توزيع.?أرباح|dividend", re.I),                     "DIVIDENDS"),
    # إيرادات
    (re.compile(r"مبيعات|إيراد.?مبيعات|revenues?.*sales?", re.I),                "SALES_REVENUE"),
    (re.compile(r"إيراد.?خدم|service.?revenue", re.I),                            "SERVICE_REVENUE"),
    (re.compile(r"إيراد.?إيجار|rental.?income|rent.?income", re.I),               "RENTAL_INCOME"),
    (re.compile(r"فائدة.?مكتسب|interest.?income", re.I),                          "INTEREST_INCOME"),
    (re.compile(r"أرباح.?استثمار|investment.?income", re.I),                      "INVESTMENT_INCOME"),
    (re.compile(r"إيراد.*عقود|إيراد.*إنشاء|contract.*revenue", re.I),             "CONTRACT_REVENUE"),
    # تكلفة مبيعات
    (re.compile(r"تكلفة.*(مبيعات|بضاعة|منتج)|COGS|cost.?of.?(sales|goods)", re.I), "COGS"),
    (re.compile(r"تكاليف.*مباشرة|direct.?cost", re.I),                            "DIRECT_COSTS"),
    # رواتب
    (re.compile(r"رواتب|أجور|مرتبات|salaries|wages", re.I),                      "SALARIES_WAGES"),
    (re.compile(r"مكافأة.*نهاية.*خدمة|EOSB|end.?of.?service", re.I),             "EOSB_EXPENSE"),
    (re.compile(r"تأمين.?(طبي|صحي)|medical.?insurance", re.I),                   "MEDICAL_INSURANCE"),
    # مصروفات تشغيلية
    (re.compile(r"إيجار.?(مدفوع|منشأة|مكتب)(?!.*مقدم)", re.I),                   "RENT_EXPENSE"),
    (re.compile(r"كهرباء|ماء|مياه|اتصالات|utilities", re.I),                      "UTILITIES"),
    (re.compile(r"إهلاك|اهلاك|استهلاك(?!.*مجمع)", re.I),                          "DEPRECIATION_EXPENSE"),
    (re.compile(r"إطفاء.?(أصل|غير.?ملموس|برنامج)", re.I),                        "AMORTIZATION_EXPENSE"),
    (re.compile(r"صيانة|إصلاح|تصليح|maintenance", re.I),                          "MAINTENANCE"),
    (re.compile(r"تسويق|دعاية|إعلان|marketing|advertising", re.I),                "MARKETING"),
    (re.compile(r"نقل|شحن|توصيل|freight|shipping", re.I),                         "FREIGHT"),
    (re.compile(r"سفر|تذاكر|بدل.?سفر|travel", re.I),                              "TRAVEL"),
    (re.compile(r"مستلزمات|قرطاسية|أدوات.*مكتب|stationery", re.I),                "STATIONERY"),
    (re.compile(r"استشار|محاسب|قانون|مهن|consulting|professional", re.I),         "PROFESSIONAL_FEES"),
    (re.compile(r"تأمين.?(ممتلكات|أصول|مسؤولية)(?!.*طبي)", re.I),                 "INSURANCE_EXPENSE"),
    # تكاليف تمويل
    (re.compile(r"فائدة.*مدفوع|فوائد.*دين|interest.?expense", re.I),              "INTEREST_EXPENSE"),
    (re.compile(r"رسوم.?بنك|عمولة.?بنك|bank.?charges", re.I),                     "BANK_CHARGES"),
    # حسابات وسيطة
    (re.compile(r"حساب.*تسوية|clearing.*account|settlement.*account", re.I),      "CLEARING_ACCOUNT"),
    (re.compile(r"معلقات|suspended|suspense", re.I),                               "BANK_SUSPENSE"),
    (re.compile(r"ض.?ق.?م.*تسوية|VAT.*settlement", re.I),                         "VAT_SETTLEMENT"),
    (re.compile(r"حساب.*مركز|head.?office", re.I),                                 "HEAD_OFFICE_ACCOUNT"),
    (re.compile(r"حساب.*فرع|branch.?account", re.I),                               "BRANCH_ACCOUNT"),
    # بنوك إسلامية (ط.1)
    (re.compile(r"مرابحة", re.I),                                                  "MURABAHA_RECEIVABLE"),
    (re.compile(r"إجارة|اجارة", re.I),                                             "IJARA_RECEIVABLE"),
    (re.compile(r"مشاركة.*استثمار", re.I),                                         "MUSHARAKA_INVESTMENT"),
    (re.compile(r"ودائع.*عملاء|customer.?deposit", re.I),                          "CUSTOMER_DEPOSITS"),
]

ENGLISH_PATTERNS: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"\bcash\b(?!.*equivalents)", re.I),                               "CASH"),
    (re.compile(r"petty.?cash|cash.?and.?cash.?equiv", re.I),                     "PETTY_CASH"),
    (re.compile(r"accounts?.?receivable|trade.?receivable", re.I),                 "ACC_RECEIVABLE"),
    (re.compile(r"notes?.?receivable", re.I),                                      "NOTES_RECEIVABLE"),
    (re.compile(r"raw.?material", re.I),                                           "RAW_MATERIALS"),
    (re.compile(r"work.?in.?progress|WIP\b", re.I),                               "WIP"),
    (re.compile(r"finished.?goods", re.I),                                         "FINISHED_GOODS"),
    (re.compile(r"inventor(?:y|ies)", re.I),                                       "INVENTORY"),
    (re.compile(r"prepaid|deferred.?expense", re.I),                               "PREPAID_EXPENSES"),
    (re.compile(r"accumulated.?depreciation|accum.?depr", re.I),                   "ACCUM_DEPR_GENERAL"),
    (re.compile(r"right.?of.?use|ROU.?asset", re.I),                               "ROU_ASSET"),
    (re.compile(r"goodwill", re.I),                                                "GOODWILL"),
    (re.compile(r"property.?plant.?equipment|PPE\b", re.I),                        "MACHINERY_EQUIPMENT"),
    (re.compile(r"accounts?.?payable|trade.?payable", re.I),                       "ACC_PAYABLE"),
    (re.compile(r"accrued.?liabilit|accrued.?expense", re.I),                      "ACCRUED_EXPENSES"),
    (re.compile(r"lease.?liabilit", re.I),                                         "LEASE_LIABILITY_NC"),
    (re.compile(r"long.?term.?(?:loan|debt)", re.I),                               "LONG_TERM_LOAN"),
    (re.compile(r"paid.?in.?capital|share.?capital", re.I),                        "PAID_IN_CAPITAL"),
    (re.compile(r"retained.?earnings", re.I),                                      "RETAINED_EARNINGS"),
    (re.compile(r"revenue|sales(?!.*return)", re.I),                               "SALES_REVENUE"),
    (re.compile(r"cost.?of.?(?:sales|goods|revenue)|COGS\b", re.I),               "COGS"),
    (re.compile(r"salar(?:y|ies)|wages|payroll", re.I),                            "SALARIES_WAGES"),
    (re.compile(r"depreciation.?expense", re.I),                                   "DEPRECIATION_EXPENSE"),
    (re.compile(r"interest.?expense", re.I),                                       "INTEREST_EXPENSE"),
    (re.compile(r"interest.?income", re.I),                                        "INTEREST_INCOME"),
    (re.compile(r"bank.?charges|bank.?fees?", re.I),                               "BANK_CHARGES"),
    (re.compile(r"vat.?input|input.?tax", re.I),                                   "VAT_INPUT"),
    (re.compile(r"vat.?output|output.?tax", re.I),                                 "VAT_OUTPUT"),
    (re.compile(r"clearing.?account|settlement.?account", re.I),                   "CLEARING_ACCOUNT"),
    (re.compile(r"suspense.?account|bank.?suspense", re.I),                        "BANK_SUSPENSE"),
]


class LexiconIndex:
    def __init__(self) -> None:
        self._exact_ar:  Dict[str, str] = {}
        self._exact_en:  Dict[str, str] = {}
        self._concepts:  Dict[str, ConceptEntry] = {}

    @classmethod
    def load(cls, excel_path: str | Path) -> "LexiconIndex":
        """يحمّل من Excel."""
        import openpyxl
        idx = cls()
        wb  = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows())]
            if "concept_id" not in headers:
                continue
            col = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    cid     = str(row[col["concept_id"]] or "").strip()
                    if not cid: continue
                    name_ar = str(row[col.get("name_ar",  -1)] or "") if "name_ar"  in col else ""
                    name_en = str(row[col.get("name_en",  -1)] or "") if "name_en"  in col else ""
                    section = str(row[col.get("section",  -1)] or "") if "section"  in col else ""
                    nature  = str(row[col.get("nature",   -1)] or "") if "nature"   in col else ""
                    level   = str(row[col.get("level",    -1)] or "") if "level"    in col else ""
                    code    = str(row[col.get("code",     -1)] or "") if "code"     in col else ""
                    ar_raw  = str(row[col.get("ar_variants",-1)] or "") if "ar_variants" in col else ""
                    en_raw  = str(row[col.get("en_variants",-1)] or "") if "en_variants" in col else ""
                    ar_vars: Set[str] = {v.strip() for v in re.split(r"[·|،,]", ar_raw) if v.strip()}
                    en_vars: Set[str] = {v.strip() for v in re.split(r"[·|،,]", en_raw) if v.strip()}
                    if name_ar: ar_vars.add(name_ar)
                    if name_en: en_vars.add(name_en)
                    idx._concepts[cid] = ConceptEntry(cid, code, name_ar, name_en, section, nature, level, ar_vars, en_vars)
                    for v in ar_vars:
                        nv = normalize_ar(v)
                        if nv and nv not in idx._exact_ar: idx._exact_ar[nv] = cid
                    for v in en_vars:
                        nv = normalize_en(v)
                        if nv and nv not in idx._exact_en: idx._exact_en[nv] = cid
                except (IndexError, TypeError):
                    continue
        wb.close()
        return idx

    @classmethod
    def from_builtin_patterns(cls) -> "LexiconIndex":
        idx = cls()
        for _, cid in ARABIC_PATTERNS + ENGLISH_PATTERNS:
            if cid not in idx._concepts:
                idx._concepts[cid] = ConceptEntry(cid,"","",cid,"","","")
        return idx

    def match(self, name: str, code: str = "") -> ConceptMatch:
        if not name:
            return ConceptMatch("UNKNOWN", 0.0, "no_match", "")
        name_str = str(name).strip()
        norm_ar  = normalize_ar(name_str)
        norm_en  = normalize_en(name_str)
        # 1. Exact Arabic — القسم 8.1: exact_match = 0.92
        if norm_ar in self._exact_ar:
            return ConceptMatch(self._exact_ar[norm_ar], 0.92, "exact_ar", norm_ar)
        # 2. Exact English — القسم 8.1: exact_match = 0.92
        if norm_en in self._exact_en:
            return ConceptMatch(self._exact_en[norm_en], 0.92, "exact_en", norm_en)
        # 3. Regex Arabic — القسم 8.1: regex_match = 0.75
        for pattern, cid in ARABIC_PATTERNS:
            if pattern.search(name_str):
                return ConceptMatch(cid, 0.75, "regex_ar", pattern.pattern[:40])
        # 4. Regex English — القسم 8.1: regex_match = 0.75
        for pattern, cid in ENGLISH_PATTERNS:
            if pattern.search(name_str):
                return ConceptMatch(cid, 0.75, "regex_en", pattern.pattern[:40])
        return ConceptMatch("UNKNOWN", 0.0, "no_match", name_str)

    def get_concept(self, cid: str) -> Optional[ConceptEntry]:
        return self._concepts.get(cid)

    def stats(self) -> Dict:
        return {
            "ar_exact":     len(self._exact_ar),
            "en_exact":     len(self._exact_en),
            "concepts":     len(self._concepts),
            "ar_patterns":  len(ARABIC_PATTERNS),
            "en_patterns":  len(ENGLISH_PATTERNS),
        }


_GLOBAL: Optional[LexiconIndex] = None

def get_lexicon(excel_path: Optional[str] = None) -> LexiconIndex:
    global _GLOBAL
    if _GLOBAL is None:
        if excel_path and Path(excel_path).exists():
            _GLOBAL = LexiconIndex.load(excel_path)
        else:
            _GLOBAL = LexiconIndex.from_builtin_patterns()
    return _GLOBAL

def reset_lexicon() -> None:
    global _GLOBAL
    _GLOBAL = None
