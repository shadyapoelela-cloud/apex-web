"""
APEX COA Engine v4.3 — Core Pipeline (100%)
============================================
ينسق الـ 8 خطوات من استقبال الملف حتى التقرير النهائي.
القسم 3 + القسم 6 + ملحق ح من الوثيقة.

الاستخدام:
    engine = COAEngine()
    result = await engine.process(file_bytes, erp_system="SAP", filename="coa.xlsx")
"""
from __future__ import annotations

import io
import json
import re
import time
import uuid as _uuid
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .config import (
    CONFIDENCE_AUTO_APPROVE, CONFIDENCE_AUTO_CLASSIFY,
    CONFIDENCE_PENDING_REVIEW, CONFIDENCE_REVIEW, CONFIDENCE_LLM_FALLBACK,
    QUALITY_WEIGHTS, QUALITY_MIN_APPROVAL, ERROR_DEDUCTIONS,
    MAX_FILE_SIZE, MAX_ROWS, ALLOWED_EXTENSIONS,
    LLM_MODEL, LLM_MAX_TOKENS, LLM_TEMPERATURE,
    LLM_TIMEOUT_SECONDS, LLM_RETRY_COUNT, LLM_RETRY_DELAY,
    LLM_FALLBACK_STATUS, ANTHROPIC_API_KEY,
    ARABIC_ENCODINGS,
)
from .error_checks import run_all_checks, summarize_errors, COAError
from .lexicon_loader import get_lexicon, normalize_ar, normalize_en, has_valid_arabic

# Wave 2 checks + advanced modules
try:
    from .error_checks_wave2 import run_wave2_checks
    HAS_WAVE2 = True
except ImportError:
    HAS_WAVE2 = False

try:
    from .advanced_checks import run_fraud_detection, detect_sector, run_cross_validation
    HAS_ADVANCED = True
except ImportError:
    HAS_ADVANCED = False

try:
    from .knowledge_graph import (
        KNOWLEDGE_GRAPH, validate_ontology,
        get_graph_context, classify_with_graph,
    )
    HAS_KNOWLEDGE_GRAPH = True
except ImportError:
    HAS_KNOWLEDGE_GRAPH = False

try:
    from .financial_simulation import (
        simulate_financial_statements,
        run_compliance_check,
        generate_implementation_roadmap,
    )
    HAS_FINANCIAL_SIM = True
except ImportError:
    HAS_FINANCIAL_SIM = False


# ─────────────────────────────────────────────────────────────
# Data Classes
# ─────────────────────────────────────────────────────────────
@dataclass
class ProcessedAccount:
    code:                  str
    name_raw:              str
    name_normalized:       str
    parent_code:           Optional[str] = None
    level_num:             int           = 1
    concept_id:            Optional[str] = None
    section:               Optional[str] = None
    nature:                str           = "unknown"
    account_level:         str           = "unknown"
    confidence:            float         = 0.0
    classification_method: str           = "unclassified"
    review_status:         str           = "pending"
    auto_fix_applied:      bool          = False
    errors:                List[str]     = field(default_factory=list)
    llm_reason:            Optional[str] = None

    def to_dict(self) -> Dict:
        return {
            "code":                  self.code,
            "name":                  self.name_raw,
            "name_raw":              self.name_raw,
            "name_normalized":       self.name_normalized,
            "parent_code":           self.parent_code,
            "level_num":             self.level_num,
            "concept_id":            self.concept_id,
            "section":               self.section,
            "nature":                self.nature,
            "account_level":         self.account_level,
            "confidence":            round(self.confidence, 3),
            "classification_method": self.classification_method,
            "review_status":         self.review_status,
            "auto_fix_applied":      self.auto_fix_applied,
            "errors":                self.errors,
        }


@dataclass
class PipelineResult:
    upload_id:          str
    file_pattern:       str
    encoding_detected:  str
    erp_system:         Optional[str]
    processing_ms:      int
    status:             str
    accounts:           List[ProcessedAccount]
    errors:             List[COAError]
    quality_score:      float
    quality_grade:      str
    quality_dimensions: Dict[str, float]
    confidence_avg:     float
    errors_summary:     Dict[str, int]
    auto_approved:      int
    pending_review:     int
    review_queue:       List[Dict]
    sector_detected:    Optional[str]
    recommendations:    List[str]
    session_health:     Dict[str, Any]
    simulation:         Optional[Dict] = None
    compliance:         Optional[Dict] = None
    roadmap:            List[Dict] = field(default_factory=list)

    def to_api_response(self) -> Dict:
        import datetime
        return {
            "upload_id":         self.upload_id,
            "processed_at":      datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "processing_ms":     self.processing_ms,
            "file_pattern":      self.file_pattern,
            "erp_system":        self.erp_system,
            "encoding_detected": self.encoding_detected,
            "status":            self.status,
            "quality_score":     round(self.quality_score, 2),
            "quality_grade":     self.quality_grade,
            "confidence_avg":    round(self.confidence_avg, 3),
            "sector_detected":   self.sector_detected,
            "quality_dimensions":self.quality_dimensions,
            "errors":            [e.to_dict() for e in self.errors],
            "errors_summary":    self.errors_summary,
            "accounts":          [a.to_dict() for a in self.accounts],
            "total_accounts":    len(self.accounts),
            "auto_approved":     self.auto_approved,
            "pending_review":    self.pending_review,
            "review_queue":      self.review_queue,
            "session_health":    self.session_health,
            "recommendations":   self.recommendations,
        }


# ─────────────────────────────────────────────────────────────
# الخطوة 0+1: قراءة الملف + كشف النمط
# ─────────────────────────────────────────────────────────────
_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def detect_encoding(raw_bytes: bytes) -> str:
    """ملحق ب — كشف ترميز CSV."""
    try:
        import chardet
        enc = chardet.detect(raw_bytes[:10_000]).get("encoding", "utf-8") or "utf-8"
        return enc
    except ImportError:
        pass
    for enc in ARABIC_ENCODINGS:
        try:
            raw_bytes[:500].decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "utf-8"


def read_file(file_bytes: bytes, filename: str) -> Tuple[List[Dict], str]:
    """يقرأ XLSX/XLS/CSV ويُعيد (rows, encoding)."""
    if not HAS_OPENPYXL:
        raise ImportError("pip install openpyxl")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "xlsx"
    encoding = "utf-8"

    if ext in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        # اختيار أفضل sheet (COA وليس journals)
        best_ws = wb.active
        best_rows = best_ws.max_row if best_ws else 0
        for ws in wb.worksheets:
            if ws.max_row > best_rows:
                headers_sample = [str(c.value or "").lower()
                                  for c in next(ws.iter_rows(max_row=1))]
                # تجاهل journals sheets
                journal_markers = {"ta", "inv", "bill", "misc", "date", "ref", "debit", "credit"}
                if not any(h in journal_markers for h in headers_sample):
                    best_ws = ws
                    best_rows = ws.max_row
        rows: List[Dict] = []
        headers = None
        for i, row in enumerate(best_ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or "").strip() for c in row]
                continue
            if all(c is None for c in row): continue
            if len(rows) >= MAX_ROWS: break
            rows.append({h: v for h, v in zip(headers or [], row)})
        wb.close()
        return rows, encoding

    elif ext == "csv":
        encoding = detect_encoding(file_bytes)
        text = file_bytes.decode(encoding, errors="replace")
        lines = text.splitlines()
        if not lines: return [], encoding
        # دعم الفواصل العربية والإنجليزية
        sep = "," if "," in lines[0] else "\t" if "\t" in lines[0] else ","
        headers = [h.strip() for h in lines[0].split(sep)]
        rows = []
        for line in lines[1:]:
            if not line.strip(): continue
            parts = line.split(sep)
            rows.append({h: (parts[i].strip() if i < len(parts) else None)
                         for i, h in enumerate(headers)})
            if len(rows) >= MAX_ROWS: break
        return rows, encoding

    raise ValueError(f"نوع الملف غير مدعوم: {ext}")


def detect_pattern(rows: List[Dict]) -> str:
    """
    كشف النمط من 12 نمط (القسم 2.1).
    مرتب من الأكثر تحديداً للأقل.
    """
    if not rows: return "UNKNOWN"
    headers = list(rows[0].keys())
    sample  = rows[:min(20, len(rows))]

    # 1. OPERATIONAL_INTEGRATED
    if len(headers) > 10:
        mixed = {"customer","vendor","product","partner","عميل","مورد","منتج","زبون"}
        if any(k in str(v).lower() for r in sample for v in r.values() for k in mixed):
            return "OPERATIONAL_INTEGRATED"

    # 2. ZOHO_BOOKS — Account ID > 15 رقم
    id_cols = [h for h in headers if "id" in h.lower() or "account id" in h.lower()]
    for col in id_cols:
        ids = [str(r.get(col,"") or "").strip() for r in sample[:5]]
        if any(len(s) > 15 and s.isdigit() for s in ids):
            return "ZOHO_BOOKS"

    # 3. MIGRATION_FILE — عمودا old/new code
    code_cols = [h for h in headers if any(k in h.lower()
                  for k in ["code","رقم","كود","رمز","account"])]
    if len(code_cols) >= 2:
        migration_kw = {"old","قديم","سابق","new","جديد","مقترح","previous","current"}
        if any(k in h.lower() for h in code_cols for k in migration_kw):
            return "MIGRATION_FILE"

    # 4. ACCOUNTS_WITH_JOURNALS — أكواد يوميات
    journal_codes = {"ta","inv","bill","misc","jnl"}
    all_codes = [str(r.get(h,"") or "") for r in sample for h in code_cols]
    if any(v.lower() in journal_codes for v in all_codes):
        return "ACCOUNTS_WITH_JOURNALS"

    # 5. HORIZONTAL_HIERARCHY — null_rate > 50%
    if headers:
        first_col = headers[0]
        null_rate = sum(1 for r in sample if r.get(first_col) is None) / max(len(sample), 1)
        if null_rate > 0.5 and len(headers) >= 3:
            return "HORIZONTAL_HIERARCHY"

    # 6. SPARSE_COLUMNAR_HIERARCHY
    level_kw = {"level","مستوى","group","مجموعة"}
    level_cols = [h for h in headers if any(k in h.lower() for k in level_kw)]
    if len(level_cols) >= 3:
        high_null = sum(
            1 for c in level_cols
            if sum(1 for r in sample if r.get(c) is None) / max(len(sample), 1) > 0.4
        )
        if high_null >= 2:
            return "SPARSE_COLUMNAR_HIERARCHY"

    # 7. HIERARCHICAL_TEXT_PARENT
    parent_cols = [h for h in headers if any(k in h.lower()
                    for k in ["parent","أب","الأب"])]
    if parent_cols:
        text_parents = [r.get(parent_cols[0]) for r in sample
                        if r.get(parent_cols[0]) and
                        not str(r.get(parent_cols[0])).replace(".","").isdigit()]
        if len(text_parents) > len(sample) * 0.3:
            return "HIERARCHICAL_TEXT_PARENT"

    # 8. HIERARCHICAL_NUMERIC_PARENT
    if parent_cols:
        num_parents = [r.get(parent_cols[0]) for r in sample
                       if r.get(parent_cols[0]) and
                       str(r.get(parent_cols[0])).replace(".","").isdigit()]
        if num_parents:
            return "HIERARCHICAL_NUMERIC_PARENT"

    # 9. ODOO_WITH_ID — __export__
    for h in headers:
        if h.lower() == "id":
            ids = [str(r.get(h,"") or "") for r in sample[:5]]
            if any("__export__" in v or "." in v[:15] for v in ids):
                return "ODOO_WITH_ID"

    # 10. ENGLISH_WITH_CLASS
    if any("class" in h.lower() for h in headers):
        return "ENGLISH_WITH_CLASS"

    # 11. ODOO_FLAT
    odoo_kw = {"user_type_id","reconcile","account_type","النوع","نوع","نوع الحساب","تصنيف الحساب"}
    if any(k in {h.lower().strip() for h in headers} for k in odoo_kw):
        return "ODOO_FLAT"

    # 12. GENERIC_FLAT
    if any(any(k in h.lower() for k in ["code","رقم","كود"]) for h in headers):
        return "GENERIC_FLAT"

    return "UNKNOWN"


# ─────────────────────────────────────────────────────────────
# الخطوة 2: توحيد الأعمدة (ملحق أ)
# ─────────────────────────────────────────────────────────────
_COL_PATTERNS: Dict[str, List[str]] = {
    "code":        [r"code|رقم.?الحساب|كود.?الحساب|رقم|الكود|الرمز|كود|رمز.?الحساب|account.?no|account.?number|account.?code"],
    "name":        [r"name|اسم.?الحساب|الاسم|وصف.?الحساب|عنوان.?الحساب|اسم|بيان|التسمية|account.?name|description"],
    "parent_code": [r"parent|أب|الأب|كود.?الأب|رمز.?الأب|حساب.?الأب|الحساب.?الرئيسي|المجموعة.?الأب|parent.?code|parent.?account"],
    "type":        [r"type|نوع|الطبيعة|تصنيف|فئة|account.?type|user.?type"],
    "nature":      [r"nature|طبيعة|balance|الطبيعة|مدين.?دائن|normal.?balance"],
    "level":       [r"level|مستوى|درجة|مرتبة"],
    "is_posting":  [r"posting|ترحيل|قيد.?مباشر|قابل.?للترحيل|reconcile"],
    "description": [r"description|وصف|ملاحظات|تعليق|شرح|notes"],
    "old_code":    [r"old.?code|الكود.?القديم|الرقم.?القديم|كود.?سابق"],
    "new_code":    [r"new.?code|الكود.?الجديد|الرقم.?الجديد|كود.?مقترح"],
}

def _norm_col(col: str) -> str:
    s = re.sub(r"[_\-\.\s]+", " ", str(col).strip().lower())
    return normalize_ar(s) or s

def map_columns(headers: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    norm_h = {_norm_col(h): h for h in headers}
    for role, patterns in _COL_PATTERNS.items():
        for pat in patterns:
            for nc, oc in norm_h.items():
                if re.search(pat, nc, re.I) and role not in mapping:
                    mapping[role] = oc; break
            if role in mapping: break

    # fallback: أعمدة الأنماط الخاصة (HORIZONTAL/SPARSE)
    if "code" not in mapping:
        for h in headers:
            if any(k in h.lower() for k in ["حساب","account","مستوى","level"]):
                mapping["code"] = h; break
    if "name" not in mapping:
        for h in headers:
            if any(k in h.lower() for k in ["اسم","name","مجموعة","group","مستوى","level"]):
                mapping["name"] = h; break

    missing = [r for r in ["code","name"] if r not in mapping]
    if missing:
        raise ValueError(f"أعمدة إلزامية غير موجودة: {missing}. الأعمدة المتاحة: {headers}")
    return mapping


def normalize_code(raw: Any) -> Optional[str]:
    """تنظيف الكود — القسم 6.3-A (10 حالات)."""
    if raw is None: return None
    s = str(raw).strip().translate(_ARABIC_DIGITS)
    if not s or s.lower() in {"nan","none","null",""}: return None
    s = re.sub(r"(\d),(\d)", r"\1\2", s)  # فاصلة آلاف
    if re.fullmatch(r"\d+\.0+", s): return str(int(float(s)))  # float trailing zeros
    if re.fullmatch(r"\d+(\.\d+)+", s): return s  # نقطة هيكلية
    try:
        if "e" in s.lower() and re.match(r"\d", s):
            return str(int(float(s)))
    except (ValueError, OverflowError): pass
    return s.strip() or None


def standardize_rows(rows: List[Dict], col_map: Dict[str, str]) -> List[Dict]:
    result = []
    for row in rows:
        code = normalize_code(row.get(col_map.get("code",""), ""))
        name = str(row.get(col_map.get("name",""), "") or "").strip()
        if not code and not name: continue
        result.append({
            "code":        code,
            "name_raw":    name,
            "parent_code": normalize_code(row.get(col_map.get("parent_code",""))),
            "type":        str(row.get(col_map.get("type",""), "") or ""),
            "nature":      str(row.get(col_map.get("nature",""), "") or ""),
            "level":       str(row.get(col_map.get("level",""), "") or ""),
            "description": str(row.get(col_map.get("description",""), "") or ""),
        })
    return result


# ─────────────────────────────────────────────────────────────
# الخطوة 3: بناء الشجرة الهرمية (القسم 3.1)
# ─────────────────────────────────────────────────────────────
def find_parent_by_prefix(code: str, code_index: Dict) -> Optional[str]:
    code = str(code).replace(".0","").strip()
    for length in range(len(code)-1, 0, -1):
        if code[:length] in code_index:
            return code[:length]
    return None


def _resolve_zoho_parents(accounts: List[Dict], code_index: Dict) -> None:
    """Zoho: parent بالاسم — قد يكون AMBIGUOUS (القسم 2.2 + 3.1)."""
    name_to_codes: Dict[str, List[str]] = {}
    for acc in accounts:
        name = str(acc.get("name_raw","") or "").strip()
        code = str(acc.get("code","") or "").strip()
        if name and code:
            name_to_codes.setdefault(name, []).append(code)

    for acc in accounts:
        pname = str(acc.get("_parent_name","") or "").strip()
        if not pname or acc.get("parent_code"):
            continue
        matches = name_to_codes.get(pname, [])
        if len(matches) == 1:
            acc["parent_code"] = matches[0]
        elif len(matches) > 1:
            acc["parent_code"] = "AMBIGUOUS"  # → مراجعة بشرية


def build_hierarchy(accounts: List[Dict]) -> Dict[str, Dict]:
    code_index: Dict[str, Dict] = {}
    for acc in accounts:
        if acc.get("code"):
            code_index[str(acc["code"])] = acc

    # المرحلة 0: حل أبوّة Zoho بالاسم (AMBIGUOUS عند التكرار)
    _resolve_zoho_parents(accounts, code_index)

    # المرحلة 1: البحث عن أب صريح موجود
    for acc in accounts:
        code = str(acc.get("code","")).strip()
        if not acc.get("parent_code") and code:
            acc["parent_code"] = find_parent_by_prefix(code, code_index)

    # المرحلة 2: بناء هرمية مُستنتجة من prefix الكود
    # إذا لم يُوجد أب صريح: 101000 → أب 10100 → أب 1010 → أب 101 → أب 10 → أب 1
    orphans = [a for a in accounts if not a.get("parent_code") and len(str(a.get("code",""))) > 1]
    if orphans:
        # اجمع كل prefixes الممكنة من كل الأكواد
        all_codes = {str(a.get("code","")).strip() for a in accounts if a.get("code")}
        prefix_set: set = set()
        for code in all_codes:
            clean = code.replace(".0","").strip()
            if clean and clean[0].isdigit():
                for length in range(1, len(clean)):
                    prefix_set.add(clean[:length])

        # أنشئ عقد وسيطة اصطناعية في code_index فقط (لا تُضاف للحسابات)
        for prefix in sorted(prefix_set):
            if prefix not in code_index:
                code_index[prefix] = {
                    "code": prefix,
                    "name_raw": f"مجموعة {prefix}",
                    "account_level": "header",
                    "_synthetic": True,
                }

        # سلسِل أبوّة العقد الاصطناعية نفسها
        for prefix in sorted(prefix_set, key=len, reverse=True):
            node = code_index.get(prefix)
            if node and node.get("_synthetic") and not node.get("parent_code"):
                node["parent_code"] = find_parent_by_prefix(prefix, code_index)

        # أعد البحث عن أب للحسابات اليتيمة
        for acc in accounts:
            code = str(acc.get("code","")).strip()
            if not acc.get("parent_code") and code:
                acc["parent_code"] = find_parent_by_prefix(code, code_index)

    def get_level(code: str, visited: set) -> int:
        if code in visited or code not in code_index: return 1
        visited.add(code)
        p = code_index[code].get("parent_code")
        if not p or p not in code_index: return 1
        return 1 + get_level(p, visited)

    for acc in accounts:
        acc["level_num"] = get_level(str(acc.get("code","")), set())

    return code_index


# ─────────────────────────────────────────────────────────────
# الخطوة 4: التصنيف (5 طبقات — القسم 6)
# ─────────────────────────────────────────────────────────────
_CODE_MAP = {
    "1":("asset","أصول","debit"),
    "2":("liability","خصوم","credit"),
    "3":("equity","حقوق الملكية","credit"),
    "4":("revenue","إيرادات","credit"),
    "5":("cogs","تكلفة المبيعات","debit"),
    "6":("expense","مصروفات","debit"),
    "7":("finance_cost","تكاليف تمويلية","debit"),
    "8":("closing","حسابات ختامية","variable"),
    "9":("verification","حسابات تحقق","variable"),
}
_SUBCODE_MAP = {
    ("1","10"):"current_asset",   ("1","11"):"current_asset",
    ("1","12"):"non_current_asset",("1","13"):"non_current_asset",
    ("1","14"):"non_current_asset",("1","15"):"non_current_asset",
    ("1","16"):"non_current_asset",("1","17"):"non_current_asset",
    ("1","18"):"non_current_asset",("1","19"):"non_current_asset",
    ("2","20"):"current_liability",("2","21"):"current_liability",
    ("2","22"):"non_current_liability",("2","23"):"non_current_liability",
    ("2","24"):"non_current_liability",("2","25"):"non_current_liability",
    ("2","26"):"non_current_liability",("2","27"):"non_current_liability",
    ("2","28"):"non_current_liability",("2","29"):"non_current_liability",
}
_USER_TYPE_MAP = {
    "بنك والنقد":         ("BANK","current_asset","debit",0.80),
    "bank and cash":      ("BANK","current_asset","debit",0.80),
    "الأصول المتداولة":   ("current_asset","current_asset","debit",0.80),
    "current assets":     ("current_asset","current_asset","debit",0.80),
    "أصول ثابتة":         ("non_current_asset","non_current_asset","debit",0.80),
    "fixed assets":       ("non_current_asset","non_current_asset","debit",0.80),
    "المدين":             ("ACC_RECEIVABLE","current_asset","debit",0.80),
    "receivable":         ("ACC_RECEIVABLE","current_asset","debit",0.80),
    "الدائن":             ("ACC_PAYABLE","current_liability","credit",0.80),
    "payable":            ("ACC_PAYABLE","current_liability","credit",0.80),
    "الالتزامات الحالية": ("current_liability","current_liability","credit",0.80),
    "current liabilities":("current_liability","current_liability","credit",0.80),
    "الالتزامات غير":     ("non_current_liability","non_current_liability","credit",0.80),
    "رأس المال":          ("equity","equity","credit",0.80),
    "equity":             ("equity","equity","credit",0.80),
    "الدخل":              ("revenue","revenue","credit",0.80),
    "income":             ("revenue","revenue","credit",0.80),
    "مصاريف":             ("expense","expense","debit",0.80),
    "expense":            ("expense","expense","debit",0.80),
    "cost of revenue":    ("cogs","cogs","debit",0.80),
    "أرباح السنة الجارية":("RETAINED_EARNINGS","equity","credit",0.80),
    "current year earnings":("RETAINED_EARNINGS","equity","credit",0.80),
}

def _layer1(code: str) -> Tuple[Optional[str], Optional[str], Optional[str], float]:
    """الطبقة 1: تصنيف من الكود."""
    if not code: return None, None, None, 0.0
    s = str(code).strip()
    first = s[0] if s else ""
    if first not in _CODE_MAP: return None, None, None, 0.0
    section, section_ar, nature = _CODE_MAP[first]
    sub = _SUBCODE_MAP.get((first, s[:2]), section)
    if first == "8": return "closing","حسابات ختامية","variable",0.85
    return sub, section_ar, nature, 0.85

def _layer2(acc: Dict) -> Tuple[Optional[str],Optional[str],Optional[str],float]:
    """الطبقة 2: user_type_id."""
    utype = str(acc.get("type","") or "").strip().lower()
    if not utype: return None, None, None, 0.0
    for key,(cid,sec,nat,conf) in _USER_TYPE_MAP.items():
        if key in utype: return cid, sec, nat, conf
    return None, None, None, 0.0

def _layer3(name: str, lexicon) -> Tuple[Optional[str],Optional[str],float]:
    """الطبقة 3: المعجم."""
    if not name: return None, None, 0.0
    m = lexicon.match(name)
    if m.concept_id == "UNKNOWN": return None, None, 0.0
    c = lexicon.get_concept(m.concept_id)
    return m.concept_id, (c.section if c else None), m.confidence

def _layer4(acc: Dict, tree: Dict) -> Tuple[Optional[str], float]:
    """الطبقة 4: كشف التعارضات — 4 أنواع (T1-T4)."""
    code = str(acc.get("code",""))
    sect = str(acc.get("section","") or "").lower()
    name = str(acc.get("name_raw","") or "")
    # T1: تعارض كود/تصنيف
    if code.startswith("1") and "liability" in sect:
        return "CODE_TYPE_CONFLICT", -0.20
    if code.startswith("2") and "asset" in sect:
        return "CODE_TYPE_CONFLICT", -0.20
    # T2: تعارض اسم/تصنيف — اسم إيراد في قسم مصروفات أو العكس
    if re.search(r"إيراد|revenue|income", name, re.I) and ("expense" in sect or "cogs" in sect):
        return "NAME_CLASS_CONFLICT", -0.20
    if re.search(r"مصروف|expense|cost", name, re.I) and "revenue" in sect:
        return "NAME_CLASS_CONFLICT", -0.20
    # T3: مجمع إهلاك مصنف كالتزام
    if re.search(r"مجمع.*(إهلاك|اهلاك)", name, re.I) and "liability" in sect:
        return "ACCUM_DEP_MISCLASS", -0.20
    return None, 0.0

async def _layer5(acc: Dict, tree: Dict) -> Tuple[Optional[str],Optional[str],Optional[str],float,Optional[str]]:
    """
    الطبقة 5: Claude API — ملحق ح.
    Returns: (concept_id, section, nature, confidence, reason)
    """
    if not ANTHROPIC_API_KEY:
        return None, None, None, 0.0, "LLM_API_KEY_MISSING"

    # بناء السياق
    parent_code = acc.get("parent_code","")
    parent_acc  = tree.get(parent_code, {})
    siblings    = [a.get("name_raw","") for c, a in tree.items()
                   if a.get("parent_code") == parent_code and c != acc.get("code","")][:4]

    prompt = (
        f"صنِّف هذا الحساب:\n"
        f"الكود: {acc.get('code','')}\n"
        f"الاسم: {acc.get('name_raw','')}\n"
        f"الأب: {parent_acc.get('name_raw','لا يوجد')} (كود: {parent_code or '—'})\n"
        f"المجاورون: {', '.join(siblings)}\n"
        f"أجِب بـ JSON:\n"
        '{"main_class":"asset|liability|equity|revenue|cogs|expense|finance_cost|closing",'
        '"sub_class":"current_asset|non_current_asset|current_liability|non_current_liability|...'
        '|operating_revenue|other_revenue|operating_expense|selling_expense|admin_expense|tax_expense",'
        '"normal_balance":"debit|credit","account_level":"header|sub|detail",'
        '"confidence":0.00,"reason":"..."}'
    )
    system_prompt = (
        "أنت محاسب قانوني خبير في IFRS والمعايير السعودية SOCPA. "
        "مهمتك تصنيف حسابات شجرة الحسابات للشركات العاملة في المملكة العربية السعودية.\n\n"
        "قواعد الإجابة الصارمة:\n"
        "1. أجِب بـ JSON فقط — لا نص قبله ولا بعده\n"
        "2. استخدم المعرّفات الثابتة التالية فقط:\n"
        "   الأقسام الرئيسية: asset|liability|equity|revenue|cogs|expense|finance_cost|closing\n"
        "   الأقسام الفرعية: current_asset|non_current_asset|current_liability|"
        "non_current_liability|operating_revenue|other_revenue|"
        "operating_expense|selling_expense|admin_expense|tax_expense\n"
        "3. الطبيعة: debit|credit\n"
        "4. المستوى: header|sub|detail\n"
        "5. الثقة: رقم 0.00-1.00\n\n"
        "لا تضف تفسيرات إضافية — JSON فقط."
    )

    VALID_MAIN = ["asset","liability","equity","revenue","cogs","expense","finance_cost","closing"]
    VALID_NATURE = ["debit","credit"]

    for attempt in range(LLM_RETRY_COUNT + 1):
        try:
            import httpx
            async with httpx.AsyncClient(timeout=LLM_TIMEOUT_SECONDS) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key":         ANTHROPIC_API_KEY,
                        "anthropic-version": "2023-06-01",
                        "content-type":      "application/json",
                    },
                    json={
                        "model":       LLM_MODEL,
                        "max_tokens":  LLM_MAX_TOKENS,
                        "temperature": LLM_TEMPERATURE,
                        "system":      system_prompt,
                        "messages":    [{"role":"user","content":prompt}],
                    },
                )
                resp.raise_for_status()
                raw = resp.json()["content"][0]["text"]

            # تحقق وتنظيف
            clean = re.sub(r"```(json)?", "", raw).strip()
            result = json.loads(clean)

            if not all(k in result for k in ["main_class","normal_balance","confidence"]):
                return None, None, None, 0.0, "LLM_MISSING_FIELDS"
            if result["main_class"] not in VALID_MAIN:
                return None, None, None, 0.0, "LLM_INVALID_MAIN_CLASS"
            if result.get("normal_balance") not in VALID_NATURE:
                result["normal_balance"] = "debit"

            conf = min(max(float(result["confidence"]), 0.0), 0.73)
            return (
                result["main_class"],
                result.get("sub_class", result["main_class"]),
                result["normal_balance"],
                conf,
                result.get("reason",""),
            )

        except Exception as e:
            if attempt < LLM_RETRY_COUNT:
                import asyncio
                await asyncio.sleep(LLM_RETRY_DELAY)
            else:
                return None, None, None, 0.0, f"LLM_ERROR: {str(e)[:50]}"

    return None, None, None, 0.0, "LLM_EXHAUSTED"


async def classify_account(acc: Dict, tree: Dict, lexicon) -> ProcessedAccount:
    """يُشغِّل كل طبقات التصنيف الـ 5."""
    code = acc.get("code","") or ""
    name = acc.get("name_raw","")

    concept_id, section, nature, conf, method = None, None, None, 0.0, "unclassified"
    llm_reason = None

    # Layer 1 — الكود
    s1, sub1, nat1, c1 = _layer1(code)
    if c1 > 0:
        concept_id, section, nature, conf, method = s1, s1, nat1, c1, "code_prefix"

    # Layer 2 — user_type
    cid2, s2, nat2, c2 = _layer2(acc)
    if c2 > conf:
        concept_id, section, nature, conf, method = cid2, s2, nat2, c2, "user_type"

    # Layer 3 — المعجم
    cid3, s3, c3 = _layer3(name, lexicon)
    if c3 > 0 and (not concept_id or c3 > conf + 0.05):
        concept_id, conf, method = cid3, c3, "lexicon"
        if s3: section = s3

    # Confidence bonuses — القسم 8.1
    # +0.10 إذا اتفق الكود مع user_type، +0.10 إذا اتفق الاسم مع الكود
    if c1 > 0 and c2 > 0 and s1 == s2:
        conf = min(1.0, conf + 0.10)
    if c1 > 0 and c3 > 0 and cid3:
        # name agrees with code section
        concept = lexicon.get_concept(cid3) if cid3 else None
        if concept and concept.section and concept.section.lower() == str(s1 or "").lower():
            conf = min(1.0, conf + 0.10)

    # Layer 4 — تعارضات
    conflict, penalty = _layer4(acc, tree)
    if conflict:
        conf = max(0.0, conf + penalty)

    # Layer 5 — Claude API (عند confidence < threshold)
    if conf < CONFIDENCE_LLM_FALLBACK:
        cid5, s5, nat5, c5, reason5 = await _layer5(acc, tree)
        if c5 > 0:
            concept_id, section, nature, conf = cid5, s5, nat5, c5
            method = "llm"
            llm_reason = reason5

    # Layer 6 — Knowledge Graph boost
    if HAS_KNOWLEDGE_GRAPH and concept_id and concept_id in KNOWLEDGE_GRAPH:
        ctx = get_graph_context(concept_id, depth=1)
        layer3_result = {"concept_id": concept_id, "section": section, "nature": nature, "confidence": conf}
        graph_result = classify_with_graph(acc, ctx, layer3_result)
        conf = graph_result.get("confidence", conf)

    # تحديد review_status — 4 شرائح (Table 57)
    if conf >= CONFIDENCE_AUTO_APPROVE:
        review_status = "auto_approved"
    elif conf >= CONFIDENCE_AUTO_CLASSIFY:
        review_status = "auto_classified"
    elif conf >= CONFIDENCE_PENDING_REVIEW:
        review_status = "pending_review"
    else:
        review_status = "rejected_pending"
    if not nature: _, _, nature, _ = _layer1(code)
    nature = nature or "unknown"

    return ProcessedAccount(
        code=str(code), name_raw=name,
        name_normalized=normalize_ar(name),
        parent_code=acc.get("parent_code"),
        level_num=acc.get("level_num",1),
        concept_id=concept_id, section=section,
        nature=nature, account_level=acc.get("account_level","unknown"),
        confidence=round(conf,3), classification_method=method,
        review_status=review_status, llm_reason=llm_reason,
    )


# ─────────────────────────────────────────────────────────────
# الخطوة 6: Quality Score (القسم 8)
# ─────────────────────────────────────────────────────────────
def compute_quality_score(accounts: List[ProcessedAccount], errors: List[COAError]) -> Tuple[float,str,Dict]:
    total = len(accounts)
    if total == 0: return 0.0, "F", {}

    classified = sum(1 for a in accounts if a.concept_id and a.confidence >= CONFIDENCE_REVIEW)
    error_score = 100.0
    for e in errors: error_score += ERROR_DEDUCTIONS.get(e.severity, 0)
    error_score = max(0.0, min(100.0, error_score)) / 100.0

    with_code = sum(1 for a in accounts if a.code and a.code.strip())
    with_name = sum(1 for a in accounts if a.name_raw and a.name_raw.strip())
    good_names = sum(1 for a in accounts if len(a.name_raw or "") > 3 and not str(a.name_raw).isdigit())
    numeric    = sum(1 for a in accounts if str(a.code).replace(".","").isdigit())
    # الاتساق = أكواد رقمية + حسابات لها هرمية (أب صريح أو مُستنتج)
    has_parent = sum(1 for a in accounts if a.parent_code)
    numeric_ratio = numeric / total
    hierarchy_ratio = has_parent / total
    # وزن مركب: 40% رقمية + 60% هرمية
    consistency = 0.4 * numeric_ratio + 0.6 * hierarchy_ratio

    dims = {
        "classification_accuracy": classified / total,
        "error_severity":          error_score,
        "completeness":            (with_code + with_name) / (total * 2),
        "naming_quality":          good_names / total,
        "code_consistency":        consistency,
    }
    score = sum(QUALITY_WEIGHTS[k] * dims[k] * 100 for k in QUALITY_WEIGHTS)
    score = round(max(0.0, min(100.0, score)), 2)
    grade = "A" if score>=90 else "B" if score>=80 else "C" if score>=70 else "D" if score>=60 else "F"
    return score, grade, {k: round(v*100,1) for k,v in dims.items()}


# ─────────────────────────────────────────────────────────────
# الخطوة 7: طابور المراجعة (القسم 3 — خطوة 7)
# ─────────────────────────────────────────────────────────────
def build_review_queue(accounts: List[ProcessedAccount], errors: List[COAError]) -> List[Dict]:
    error_map:    Dict[str,List[str]] = {}
    critical_set: set = set()
    for e in errors:
        if e.account_code:
            error_map.setdefault(e.account_code,[]).append(e.error_code)
            if e.severity == "Critical": critical_set.add(e.account_code)

    queue = []
    for a in accounts:
        in_queue = (a.confidence < CONFIDENCE_REVIEW or a.code in critical_set
                    or a.review_status in ("pending_review", "rejected_pending"))
        if in_queue:
            reasons = []
            if a.confidence < CONFIDENCE_REVIEW: reasons.append(f"confidence={a.confidence:.2f} < {CONFIDENCE_REVIEW}")
            if a.code in critical_set: reasons.append("خطأ Critical مفتوح")
            if a.review_status == "rejected_pending": reasons.append("ثقة منعدمة < 50%")
            queue.append({
                "account_code": a.code,
                "account_name": a.name_raw,
                "confidence":   a.confidence,
                "reason":       " + ".join(reasons),
                "error_codes":  error_map.get(a.code,[]),
                "suggested_fix":None,
            })
    return sorted(queue, key=lambda x: x["confidence"])


# ─────────────────────────────────────────────────────────────
# Auto-Fix Logic (E01, E07, E08, E09, E10, E17-E20)
# ─────────────────────────────────────────────────────────────
_SECTION_FOR_CODE = {
    "1": "asset", "2": "liability", "3": "equity",
    "4": "revenue", "5": "cogs", "6": "expense", "7": "finance_cost",
}
_NATURE_FOR_SECTION = {
    "asset": "debit", "current_asset": "debit", "non_current_asset": "debit",
    "liability": "credit", "current_liability": "credit", "non_current_liability": "credit",
    "equity": "credit", "revenue": "credit", "cogs": "debit",
    "expense": "debit", "finance_cost": "debit",
}


def _apply_auto_fix(
    error: COAError,
    acc_dicts: List[Dict],
    processed: List["ProcessedAccount"],
) -> Optional[Dict]:
    """
    يُطبِّق الإصلاح التلقائي ويُسجِّل التغيير.
    Returns dict with fix details for auto_fix_log, or None if no fix applied.
    """
    code = error.account_code
    if not code:
        return None

    # ابحث عن الحساب في القائمة
    pa = next((p for p in processed if p.code == code), None)
    if not pa:
        return None

    fix_type = None
    before_value = ""
    after_value = ""

    if error.error_code == "E01":
        # E01: duplicate code — append suffix to deduplicate
        dup_count = sum(1 for p in processed if p.code == code)
        if dup_count > 1:
            # Fix: mark subsequent duplicates
            seen = False
            for p in processed:
                if p.code == code:
                    if seen:
                        old_code = p.code
                        p.code = f"{code}_{dup_count}"
                        fix_type = "deduplicate_code"
                        before_value = old_code
                        after_value = p.code
                        p.auto_fix_applied = True
                        break
                    seen = True

    elif error.error_code == "E07":
        # E07: mix parent-child — set account_level correctly
        has_children = any(p.parent_code == code for p in processed)
        old_level = pa.account_level
        pa.account_level = "header" if has_children else "detail"
        if old_level != pa.account_level:
            fix_type = "fix_account_level"
            before_value = old_level
            after_value = pa.account_level
            pa.auto_fix_applied = True

    elif error.error_code == "E08":
        # E08: broken hierarchy — try prefix matching
        from_parent = pa.parent_code
        code_index = {p.code: p for p in processed}
        if pa.parent_code and pa.parent_code not in code_index:
            # Try prefix matching
            for length in range(len(code) - 1, 0, -1):
                prefix = code[:length]
                if prefix in code_index:
                    pa.parent_code = prefix
                    fix_type = "fix_parent_code"
                    before_value = from_parent or ""
                    after_value = prefix
                    pa.auto_fix_applied = True
                    break

    elif error.error_code == "E09":
        # E09: asset classified as liability — fix section
        first = code[0] if code else ""
        correct_section = _SECTION_FOR_CODE.get(first)
        if correct_section and pa.section and "liability" in pa.section.lower() and first == "1":
            old_section = pa.section
            pa.section = "current_asset" if len(code) >= 2 and code[1] in "01" else "non_current_asset"
            fix_type = "fix_section"
            before_value = old_section
            after_value = pa.section
            pa.auto_fix_applied = True

    elif error.error_code == "E10":
        # E10: revenue as expense — fix section
        if code.startswith("4") and pa.section and "expense" in pa.section.lower():
            old_section = pa.section
            pa.section = "revenue"
            pa.nature = "credit"
            fix_type = "fix_section"
            before_value = old_section
            after_value = "revenue"
            pa.auto_fix_applied = True
        elif code.startswith(("5", "6")) and pa.section and "revenue" in pa.section.lower():
            old_section = pa.section
            pa.section = "expense"
            pa.nature = "debit"
            fix_type = "fix_section"
            before_value = old_section
            after_value = "expense"
            pa.auto_fix_applied = True

    elif error.error_code in ("E17", "E18", "E19", "E20"):
        # E17-E20: reversed nature — fix nature based on section
        correct_nature = _NATURE_FOR_SECTION.get(pa.section or "", None)
        if correct_nature and pa.nature != correct_nature:
            # مخصصات طبيعتها دائنة حتى لو تحت الأصول
            name = pa.name_raw or ""
            if re.search(r"مخصص|مجمع|provision|accumulated|contra", name, re.I):
                correct_nature = "credit"
            old_nature = pa.nature
            pa.nature = correct_nature
            fix_type = "fix_nature"
            before_value = old_nature
            after_value = correct_nature
            pa.auto_fix_applied = True

    if fix_type:
        return {
            "account_code": code,
            "fix_type": fix_type,
            "error_code": error.error_code,
            "before_value": before_value,
            "after_value": after_value,
            "fix_confidence": pa.confidence,
            "fix_reason_ar": error.suggestion_ar,
        }
    return None


# ─────────────────────────────────────────────────────────────
# المحرك الرئيسي
# ─────────────────────────────────────────────────────────────
class COAEngine:
    def __init__(self, lexicon_path: Optional[str] = None) -> None:
        self._lexicon = get_lexicon(lexicon_path)

    async def process(
        self,
        file_bytes: bytes,
        erp_system: Optional[str] = None,
        filename:   str = "upload.xlsx",
        upload_id:  Optional[str] = None,
    ) -> PipelineResult:
        t0 = time.time()
        upload_id = upload_id or str(_uuid.uuid4())

        # 0. تحقق أولي
        if len(file_bytes) > MAX_FILE_SIZE:
            raise ValueError(f"حجم الملف {len(file_bytes):,} يتجاوز الحد {MAX_FILE_SIZE:,}")
        ext = "." + (filename.rsplit(".",1)[-1].lower() if "." in filename else "xlsx")
        if ext not in ALLOWED_EXTENSIONS:
            raise ValueError(f"نوع الملف {ext} غير مدعوم")

        # 1. قراءة + كشف النمط
        rows, encoding = read_file(file_bytes, filename)
        if not rows: raise ValueError("الملف فارغ")
        pattern = detect_pattern(rows)

        # EC5: رفض OPERATIONAL_INTEGRATED
        if pattern == "OPERATIONAL_INTEGRATED":
            return PipelineResult(
                upload_id=upload_id, file_pattern=pattern,
                encoding_detected=encoding, erp_system=erp_system,
                processing_ms=int((time.time()-t0)*1000),
                status="rejected", accounts=[], errors=[],
                quality_score=0.0, quality_grade="F", quality_dimensions={},
                confidence_avg=0.0,
                errors_summary={"critical":1,"high":0,"medium":0,"low":0,"total":1},
                auto_approved=0, pending_review=0, review_queue=[],
                sector_detected=None, recommendations=[
                    "الملف يخلط بيانات COA مع بيانات تشغيلية (عملاء/موردين) — ارفع ملف COA منفصلاً"
                ],
                session_health={"pass_one_rate":0,"pass_two_rate":0,"llm_rate":0},
            )

        # 2. توحيد الأعمدة
        col_map  = map_columns(list(rows[0].keys()))
        std_rows = standardize_rows(rows, col_map)
        if not std_rows: raise ValueError("لم يُستخرج أي حساب بعد توحيد الأعمدة")

        # 3. بناء الشجرة
        tree = build_hierarchy(std_rows)

        # 4. التصنيف (async لدعم الطبقة 5)
        processed: List[ProcessedAccount] = []
        pass1_count = pass2_count = llm_count = 0
        for acc in std_rows:
            pa = await classify_account(acc, tree, self._lexicon)
            processed.append(pa)
            if pa.classification_method in ("code_prefix","user_type"): pass1_count += 1
            elif pa.classification_method == "llm": llm_count += 1
            else: pass2_count += 1

        # 5. فحص الأخطاء — Wave 1 + Wave 2 + Advanced
        acc_dicts = [a.to_dict() for a in processed]
        errors    = run_all_checks(acc_dicts, tree)

        # Wave 2 checks (E21-E50 + EP1-EP3 + EC1-EC5)
        sector_detected = None
        if HAS_ADVANCED:
            sector_detected = detect_sector(acc_dicts)

        if HAS_WAVE2:
            w2_errors = run_wave2_checks(
                acc_dicts, tree,
                erp_system=erp_system,
                sector=sector_detected,
            )
            errors.extend(w2_errors)

        # Cross-validation (ملحق و)
        if HAS_ADVANCED:
            cv_errors = run_cross_validation(acc_dicts)
            errors.extend(cv_errors)

        # Fraud detection (ملحق س)
        if HAS_ADVANCED:
            fraud_alerts = run_fraud_detection(acc_dicts)
            for fa in fraud_alerts:
                errors.append(fa.to_error())

            # FP08: Partner drawings detection (ملحق ك-2)
            from .advanced_checks import detect_fp08_partner_drawings
            fp08 = detect_fp08_partner_drawings(acc_dicts)
            if fp08:
                errors.append(fp08.to_error())

        # Ontology validation (ملحق خ — Knowledge Graph)
        if HAS_KNOWLEDGE_GRAPH:
            ontology_errors = validate_ontology(acc_dicts)
            for oe in ontology_errors:
                errors.append(COAError(
                    error_code=oe.get("error_code", "E28"),
                    severity=oe.get("severity", "Medium"),
                    category=oe.get("category", "ontology"),
                    account_code=oe.get("account_code"),
                    account_name=oe.get("account_name"),
                    description_ar=oe.get("description_ar", ""),
                    cause_ar=oe.get("cause_ar", ""),
                    suggestion_ar=oe.get("suggestion_ar", ""),
                    auto_fixable=oe.get("auto_fixable", False),
                    references=oe.get("references", []),
                ))

        # 5c. Financial Simulation & Compliance (ملحق ن + ل2)
        simulation_result = None
        compliance_result = None
        roadmap_result = []
        if HAS_FINANCIAL_SIM:
            simulation_result = simulate_financial_statements(acc_dicts)
            compliance_result = run_compliance_check(acc_dicts, sector_detected)
            # Recommendations will be added after recs is initialized (below)

        # 5b. Auto-Fix — تطبيق الإصلاح التلقائي للأخطاء القابلة
        auto_fix_log_entries: List[Dict] = []
        for e in errors:
            if not e.auto_fixable or e.auto_fix_applied:
                continue
            fixed = _apply_auto_fix(e, acc_dicts, processed)
            if fixed:
                e.auto_fix_applied = True
                auto_fix_log_entries.append(fixed)

        # ربط الأخطاء بالحسابات
        err_by_code: Dict[str,List[str]] = {}
        crit_codes: set = set()
        for e in errors:
            if e.account_code:
                err_by_code.setdefault(e.account_code,[]).append(e.error_code)
                if e.severity == "Critical": crit_codes.add(e.account_code)
        for pa in processed:
            pa.errors = err_by_code.get(pa.code,[])
            if pa.code in crit_codes: pa.review_status = "blocked"

        # 6. التقرير
        quality_score, quality_grade, quality_dims = compute_quality_score(processed, errors)
        errors_summary = summarize_errors(errors)
        confidence_avg = sum(a.confidence for a in processed)/len(processed) if processed else 0.0

        # إحصاءات 4-tier
        auto_approved   = sum(1 for a in processed if a.review_status == "auto_approved")
        auto_classified = sum(1 for a in processed if a.review_status == "auto_classified")
        pending_rev     = sum(1 for a in processed if a.review_status in ("pending_review","pending","rejected_pending"))

        # 7. طابور المراجعة
        review_queue = build_review_queue(processed, errors)

        # الحالة النهائية
        if errors_summary["critical"] > 0:   status = "blocked"
        elif pending_rev > 0:                status = "pending_review"
        elif quality_score >= QUALITY_MIN_APPROVAL: status = "classified"
        else:                                status = "pending_review"

        # توصيات
        recs = []
        if errors_summary["critical"]: recs.append(f"{errors_summary['critical']} خطأ Critical — لا اعتماد حتى الحل")
        if confidence_avg < 0.75:      recs.append("متوسط الثقة منخفض — راجع جودة البيانات")
        if quality_score < QUALITY_MIN_APPROVAL: recs.append(f"درجة الجودة {quality_score:.1f} أقل من الحد {QUALITY_MIN_APPROVAL}")

        # Add simulation/compliance recommendations
        if HAS_FINANCIAL_SIM and simulation_result:
            for gap in simulation_result.get("structural_gaps", []):
                if gap.get("severity") in ("Critical", "High"):
                    fix = gap.get("fix", gap.get("message_ar", ""))
                    if fix and fix not in recs:
                        recs.append(fix)
        if HAS_FINANCIAL_SIM and compliance_result:
            for fail in compliance_result.get("failed", []):
                if fail.get("severity") == "Critical":
                    req = fail.get("requirement_ar", "")
                    if req and req not in recs:
                        recs.append(req)

        total = len(processed)
        session_health = {
            "pass_one_rate":    round(pass1_count/total, 3) if total else 0,
            "pass_two_rate":    round(pass2_count/total, 3) if total else 0,
            "llm_rate":         round(llm_count/total, 3)  if total else 0,
            "auto_fix_count":   len(auto_fix_log_entries),
            "auto_classified":  auto_classified,
        }

        # Generate roadmap (after all errors finalized)
        if HAS_FINANCIAL_SIM and simulation_result and compliance_result:
            roadmap_result = generate_implementation_roadmap(
                errors, simulation_result, compliance_result,
            )

        return PipelineResult(
            upload_id=upload_id, file_pattern=pattern,
            encoding_detected=encoding, erp_system=erp_system,
            processing_ms=int((time.time()-t0)*1000),
            status=status, accounts=processed, errors=errors,
            quality_score=quality_score, quality_grade=quality_grade,
            quality_dimensions=quality_dims, confidence_avg=round(confidence_avg,3),
            errors_summary=errors_summary, auto_approved=auto_approved,
            pending_review=pending_rev, review_queue=review_queue,
            sector_detected=sector_detected, recommendations=recs,
            session_health=session_health,
            simulation=simulation_result,
            compliance=compliance_result,
            roadmap=roadmap_result,
        )
