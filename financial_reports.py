import io
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dataclasses import dataclass

# تسجيل الخط العربي
_font_registered = False
def _get_arabic_font():
    global _font_registered
    if not _font_registered:
        font_path = os.path.join(os.path.dirname(__file__), "Amiri-Regular.ttf")
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("Arabic", font_path))
            _font_registered = True
            return "Arabic"
    return "Arabic" if _font_registered else "Helvetica"

def ar(text):
    """تحويل النص العربي للعرض الصحيح في PDF"""
    if not text or not isinstance(text, str):
        return str(text) if text else ""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text


@dataclass
class FinancialStatements:
    company_name: str
    period: str
    currency: str
    revenue: float
    cogs: float
    gross_profit: float
    gross_margin: float
    operating_expenses: float
    operating_profit: float
    interest_expense: float
    net_profit: float
    net_margin: float
    fixed_assets: float
    current_assets: float
    cash: float
    receivables: float
    inventory: float
    total_assets: float
    current_liabilities: float
    long_term_liabilities: float
    total_liabilities: float
    equity: float
    total_liabilities_equity: float
    current_ratio: float
    debt_to_equity: float
    roe: float
    roa: float


def build_statements(data) -> FinancialStatements:
    d = data
    revenue = abs(d.revenue)
    cogs = abs(d.cost_of_goods_sold)
    gross_profit = revenue - cogs
    gross_margin = (gross_profit / revenue * 100) if revenue else 0
    opex = abs(d.operating_expenses)
    operating_profit = gross_profit - opex
    interest = abs(d.interest_expense)
    net_profit = operating_profit - interest
    net_margin = (net_profit / revenue * 100) if revenue else 0
    total_assets = abs(d.total_assets)
    current_assets = abs(d.current_assets)
    fixed_assets = total_assets - current_assets
    cash = abs(d.cash)
    inventory = abs(d.inventory)
    receivables = max(current_assets - cash - inventory, 0)
    total_liabilities = abs(d.total_liabilities)
    current_liabilities = abs(d.current_liabilities)
    long_term = max(total_liabilities - current_liabilities, 0)
    equity = abs(d.equity)
    current_ratio = (current_assets / current_liabilities) if current_liabilities else 0
    dte = (total_liabilities / equity) if equity else 0
    roe = (net_profit / equity * 100) if equity else 0
    roa = (net_profit / total_assets * 100) if total_assets else 0
    return FinancialStatements(
        company_name="شركة أبيكس", period="2025", currency="ريال سعودي",
        revenue=revenue, cogs=cogs, gross_profit=gross_profit, gross_margin=gross_margin,
        operating_expenses=opex, operating_profit=operating_profit,
        interest_expense=interest, net_profit=net_profit, net_margin=net_margin,
        fixed_assets=fixed_assets, current_assets=current_assets, cash=cash,
        receivables=receivables, inventory=inventory, total_assets=total_assets,
        current_liabilities=current_liabilities, long_term_liabilities=long_term,
        total_liabilities=total_liabilities, equity=equity,
        total_liabilities_equity=total_liabilities + equity,
        current_ratio=current_ratio, debt_to_equity=dte, roe=roe, roa=roa)


def fmt(n):
    if isinstance(n, str): return n
    return f"({abs(n):,.0f})" if n < 0 else f"{n:,.0f}"


def generate_pdf(fs: FinancialStatements) -> bytes:
    FONT = _get_arabic_font()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    GOLD = colors.HexColor('#C9A84C')
    NAVY = colors.HexColor('#050D1A')
    NAVY2 = colors.HexColor('#0D1829')
    WHITE = colors.white
    GRAY = colors.HexColor('#8A8880')
    GREEN = colors.HexColor('#2ECC8A')
    RED = colors.HexColor('#E05050')

    title_style = ParagraphStyle('T', fontName=FONT, fontSize=20,
        textColor=GOLD, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('S', fontName=FONT, fontSize=11,
        textColor=GRAY, alignment=TA_CENTER, spaceAfter=2)
    sec_style = ParagraphStyle('SC', fontName=FONT, fontSize=13,
        textColor=WHITE, alignment=TA_RIGHT, spaceBefore=12, spaceAfter=6)

    elements = []
    elements.append(Paragraph("APEX", title_style))
    elements.append(Paragraph(ar("القوائم المالية") + " - " + fs.period, sub_style))
    elements.append(Paragraph(ar("العملة: ريال سعودي"), sub_style))
    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=GOLD))
    elements.append(Spacer(1, 0.4*cm))

    W = A4[0] - 3*cm

    def tbl(data, widths):
        t = Table(data, colWidths=widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY2),
            ('TEXTCOLOR', (0,0), (-1,0), GOLD),
            ('FONTNAME', (0,0), (-1,-1), FONT),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BACKGROUND', (0,1), (-1,-1), NAVY),
            ('TEXTCOLOR', (0,1), (-1,-1), WHITE),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [NAVY, NAVY2]),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#1E2640')),
            ('ALIGN', (0,0), (0,-1), 'RIGHT'),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]))
        return t

    # قائمة الدخل
    elements.append(Paragraph(ar("قائمة الدخل"), sec_style))
    t1 = tbl([
        [ar("المبلغ"), ar("البند")],
        [fmt(fs.revenue), ar("اجمالي المبيعات")],
        [fmt(fs.cogs), ar("تكلفة البضاعة المباعة")],
        [fmt(fs.gross_profit), ar("مجمل الربح")],
        [f'{fs.gross_margin:.1f}%', ar("هامش مجمل الربح")],
        [fmt(fs.operating_expenses), ar("المصروفات التشغيلية")],
        [fmt(fs.operating_profit), ar("الربح التشغيلي")],
        [fmt(fs.interest_expense), ar("مصروف الفوائد")],
        [fmt(fs.net_profit), ar("صافي الربح / الخسارة")],
        [f'{fs.net_margin:.1f}%', ar("هامش صافي الربح")],
    ], [W*0.35, W*0.65])
    t1.setStyle(TableStyle([
        ('TEXTCOLOR', (0,3), (0,3), GREEN if fs.gross_profit >= 0 else RED),
        ('TEXTCOLOR', (0,8), (0,8), GREEN if fs.net_profit >= 0 else RED),
        ('FONTNAME', (0,8), (0,8), FONT),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 0.5*cm))

    # الميزانية - الاصول
    elements.append(Paragraph(ar("الميزانية العمومية"), sec_style))
    t2 = tbl([
        [ar("المبلغ"), ar("الاصول")],
        [fmt(fs.fixed_assets), ar("الاصول الثابتة")],
        [fmt(fs.cash), ar("النقدية وما يعادلها")],
        [fmt(fs.receivables), ar("ذمم مدينة")],
        [fmt(fs.inventory), ar("المخزون")],
        [fmt(fs.current_assets), ar("اجمالي الاصول المتداولة")],
        [fmt(fs.total_assets), ar("اجمالي الاصول")],
    ], [W*0.35, W*0.65])
    t2.setStyle(TableStyle([
        ('FONTNAME', (0,6), (-1,6), FONT),
        ('TEXTCOLOR', (0,6), (0,6), GOLD),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 0.3*cm))

    # الميزانية - الخصوم
    t3 = tbl([
        [ar("المبلغ"), ar("الخصوم وحقوق الملكية")],
        [fmt(fs.current_liabilities), ar("الالتزامات المتداولة")],
        [fmt(fs.long_term_liabilities), ar("الالتزامات طويلة الاجل")],
        [fmt(fs.total_liabilities), ar("اجمالي الالتزامات")],
        [fmt(fs.equity), ar("حقوق الملكية")],
        [fmt(fs.total_liabilities_equity), ar("اجمالي الخصوم وحقوق الملكية")],
    ], [W*0.35, W*0.65])
    t3.setStyle(TableStyle([
        ('FONTNAME', (0,5), (-1,5), FONT),
        ('TEXTCOLOR', (0,5), (0,5), GOLD),
    ]))
    elements.append(t3)
    elements.append(Spacer(1, 0.5*cm))

    # النسب المالية
    elements.append(Paragraph(ar("النسب المالية الرئيسية"), sec_style))
    elements.append(tbl([
        [ar("القيمة"), ar("النسبة"), ar("الفئة")],
        [f'{fs.current_ratio:.2f}x', ar("نسبة السيولة الجارية"), ar("السيولة")],
        [f'{fs.debt_to_equity:.2f}x', ar("نسبة الدين / حقوق الملكية"), ar("الرفع المالي")],
        [f'{fs.roe:.1f}%', ar("العائد على حقوق الملكية ROE"), ar("الربحية")],
        [f'{fs.roa:.1f}%', ar("العائد على الاصول ROA"), ar("الربحية")],
        [f'{fs.gross_margin:.1f}%', ar("هامش الربح الاجمالي"), ar("الربحية")],
        [f'{fs.net_margin:.1f}%', ar("هامش صافي الربح"), ar("الربحية")],
    ], [W*0.2, W*0.5, W*0.3]))

    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=GOLD))
    footer = ParagraphStyle('F', fontName=FONT, fontSize=8, textColor=GRAY, alignment=TA_CENTER)
    elements.append(Paragraph(f"APEX Financial Advisory | {fs.period}", footer))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel(fs: FinancialStatements) -> bytes:
    wb = openpyxl.Workbook()
    GOLD='C9A84C'; NAVY='050D1A'; NAVY2='0D1829'; WHITE='F0EDE6'; GREEN='2ECC8A'; RED='E05050'

    def hdr(ws, r, c, v):
        cell = ws.cell(r, c, v)
        cell.font = Font(name='Arial', bold=True, color=GOLD, size=11)
        cell.fill = PatternFill('solid', fgColor=NAVY2)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='1E2640'),
            right=Side(style='thin', color='1E2640'),
            top=Side(style='thin', color='1E2640'),
            bottom=Side(style='thin', color='1E2640'))

    def row(ws, r, c, v, bold=False, color=None):
        cell = ws.cell(r, c, v)
        cell.font = Font(name='Arial', bold=bold, color=color or WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=NAVY if r % 2 == 0 else NAVY2)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='1E2640'),
            right=Side(style='thin', color='1E2640'),
            top=Side(style='thin', color='1E2640'),
            bottom=Side(style='thin', color='1E2640'))

    def setup_sheet(ws, title_text, cols=2):
        merge = f'A1:{chr(64+cols)}1'
        ws.merge_cells(merge)
        c = ws['A1']
        c.value = title_text
        c.font = Font(name='Arial', bold=True, color=GOLD, size=16)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        ws.sheet_view.rightToLeft = True
        for i in range(1, cols+1):
            ws.column_dimensions[chr(64+i)].width = 35 if i == 1 else 22

    # قائمة الدخل
    ws1 = wb.active
    ws1.title = 'قائمة الدخل'
    setup_sheet(ws1, f'قائمة الدخل - {fs.period}')
    hdr(ws1, 2, 1, 'البند')
    hdr(ws1, 2, 2, f'المبلغ ({fs.currency})')
    for i, (l, v, b, c) in enumerate([
        ('اجمالي المبيعات', fmt(fs.revenue), False, GOLD),
        ('تكلفة البضاعة المباعة', fmt(fs.cogs), False, WHITE),
        ('مجمل الربح', fmt(fs.gross_profit), True, GREEN if fs.gross_profit >= 0 else RED),
        ('هامش مجمل الربح', f'{fs.gross_margin:.1f}%', False, WHITE),
        ('المصروفات التشغيلية', fmt(fs.operating_expenses), False, WHITE),
        ('الربح التشغيلي', fmt(fs.operating_profit), True, GREEN if fs.operating_profit >= 0 else RED),
        ('مصروف الفوائد', fmt(fs.interest_expense), False, WHITE),
        ('صافي الربح / الخسارة', fmt(fs.net_profit), True, GREEN if fs.net_profit >= 0 else RED),
        ('هامش صافي الربح', f'{fs.net_margin:.1f}%', False, WHITE),
    ], 3):
        row(ws1, i, 1, l, b, c)
        row(ws1, i, 2, v, b, c)

    # الميزانية العمومية
    ws2 = wb.create_sheet('الميزانية العمومية')
    setup_sheet(ws2, f'الميزانية العمومية - {fs.period}')
    hdr(ws2, 2, 1, 'البند')
    hdr(ws2, 2, 2, f'المبلغ ({fs.currency})')
    for i, (l, v, b, c) in enumerate([
        ('الاصول الثابتة', fmt(fs.fixed_assets), False, WHITE),
        ('النقدية وما يعادلها', fmt(fs.cash), False, WHITE),
        ('ذمم مدينة', fmt(fs.receivables), False, WHITE),
        ('المخزون', fmt(fs.inventory), False, WHITE),
        ('اجمالي الاصول المتداولة', fmt(fs.current_assets), True, GOLD),
        ('اجمالي الاصول', fmt(fs.total_assets), True, GOLD),
        ('الالتزامات المتداولة', fmt(fs.current_liabilities), False, WHITE),
        ('الالتزامات طويلة الاجل', fmt(fs.long_term_liabilities), False, WHITE),
        ('اجمالي الالتزامات', fmt(fs.total_liabilities), True, WHITE),
        ('حقوق الملكية', fmt(fs.equity), False, WHITE),
        ('اجمالي الخصوم وحقوق الملكية', fmt(fs.total_liabilities_equity), True, GOLD),
    ], 3):
        row(ws2, i, 1, l, b, c)
        row(ws2, i, 2, v, b, c)

    # النسب المالية
    ws3 = wb.create_sheet('النسب المالية')
    setup_sheet(ws3, f'النسب المالية - {fs.period}', cols=3)
    ws3.column_dimensions['C'].width = 20
    hdr(ws3, 2, 1, 'النسبة')
    hdr(ws3, 2, 2, 'القيمة')
    hdr(ws3, 2, 3, 'الفئة')
    for i, (l, v, cat) in enumerate([
        ('نسبة السيولة الجارية', f'{fs.current_ratio:.2f}x', 'السيولة'),
        ('نسبة الدين / حقوق الملكية', f'{fs.debt_to_equity:.2f}x', 'الرفع المالي'),
        ('العائد على حقوق الملكية ROE', f'{fs.roe:.1f}%', 'الربحية'),
        ('العائد على الاصول ROA', f'{fs.roa:.1f}%', 'الربحية'),
        ('هامش الربح الاجمالي', f'{fs.gross_margin:.1f}%', 'الربحية'),
        ('هامش صافي الربح', f'{fs.net_margin:.1f}%', 'الربحية'),
    ], 3):
        row(ws3, i, 1, l)
        row(ws3, i, 2, v, color=GREEN)
        row(ws3, i, 3, cat)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
